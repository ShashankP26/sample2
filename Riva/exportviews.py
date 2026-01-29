from collections import defaultdict
import csv
import json
import os
from django.conf import settings
from django.http import HttpResponse
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from io import BytesIO
from django.shortcuts import get_object_or_404
from .models import Enquiry

# CSV Export View
def export_confirmed_orders_csv(request):
    # Get confirmed orders
    confirmed_orders = Enquiry.objects.filter(is_confirmed=True).prefetch_related('confirmed_enquiry_set')

    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="confirmed_orders.csv"'

    writer = csv.writer(response)
    # Write header
    writer.writerow(['#', 'Company Name', 'Customer Name', 'Contact', 'Product', 'ID', 'Closing Date', 'Quotation Number'])

    # Write data
    for index, order in enumerate(confirmed_orders, start=1):
        for confirmed_entry in order.confirmed_enquiry_set.all():
            writer.writerow([
                index,
                order.companyname,
                order.customername,
                order.contact,
                order.products,
                order.id,
                order.closuredate,
                confirmed_entry.quotation if not confirmed_entry.relegate else "Relegated",
            ])
    return response


from django.http import HttpResponse
from openpyxl import Workbook

def export_confirmed_orders_xlsx(request):
    # Fetch data
    current_user = request.user
    if current_user.is_superuser:
        confirmed_orders = Enquiry.objects.filter(is_confirmed=True).prefetch_related('confirmed_enquiry_set', 'products')
    else:
        confirmed_orders = Enquiry.objects.filter(is_confirmed=True, created_by=current_user).prefetch_related('confirmed_enquiry_set', 'products')

    # Create an Excel workbook and worksheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Confirmed Orders"

    # Write header row
    headers = ["#", "Company Name", "Customer Name", "Contact", "Product", "ID", "Closing Date", "Quotation"]
    worksheet.append(headers)

    # Write data rows
    for index, order in enumerate(confirmed_orders, start=1):
        for confirmed_entry in order.confirmed_enquiry_set.all():
            # Convert related fields to strings
            product_name = str(order.products) if order.products else ""  # Handle related `Products` object
            closing_date = order.closuredate.strftime("%Y-%m-%d") if order.closuredate else ""
            
            row = [
                index,  # Serial number
                order.companyname or "",  # Company name
                order.customername or "",  # Customer name
                str(order.contact) if order.contact else "",  # Contact
                product_name,  # Product name
                order.id,  # Order ID
                closing_date,  # Closing date
                confirmed_entry.quotation or "",  # Quotation
            ]
            worksheet.append(row)

    # Prepare the HTTP response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="confirmed_orders.xlsx"'

    # Save the workbook to the response
    workbook.save(response)
    return response



from django.http import HttpResponse
from reportlab.pdfgen import canvas
from django.http import HttpResponse
from reportlab.pdfgen import canvas

def export_confirmed_orders_pdf(request):
    # Fetch data
    current_user = request.user
    if current_user.is_superuser:
        confirmed_orders = Enquiry.objects.filter(is_confirmed=True).prefetch_related('confirmed_enquiry_set', 'products')
    else:
        confirmed_orders = Enquiry.objects.filter(is_confirmed=True, created_by=current_user).prefetch_related('confirmed_enquiry_set', 'products')

    # Create a PDF response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="confirmed_orders.pdf"'

    # Start writing PDF content
    pdf_canvas = canvas.Canvas(response)
    pdf_canvas.setFont("Helvetica", 12)

    y = 800  # Start position for text

    # Write headers
    headers = ["#", "Company Name", "Cust Name", "Contact", "Product", "ID", "Closing Date", "Quotation"]
    pdf_canvas.drawString(50, y, "Confirmed Orders")
    y -= 20

    pdf_canvas.setFont("Helvetica-Bold", 10)
    for i, header in enumerate(headers):
        pdf_canvas.drawString(50 + (i * 100), y, header)
    y -= 20

    # Write rows
    pdf_canvas.setFont("Helvetica", 10)
    for index, order in enumerate(confirmed_orders, start=1):
        for confirmed_entry in order.confirmed_enquiry_set.all():
            pdf_canvas.drawString(50, y, str(index))  # Row number
            pdf_canvas.drawString(150, y, order.companyname or "")  # Company Name
            pdf_canvas.drawString(250, y, order.customername or "")  # Customer Name
            pdf_canvas.drawString(350, y, str(order.contact))  # Contact (explicit string conversion)
            pdf_canvas.drawString(450, y, str(order.products) if order.products else "")  # Product (as string)
            pdf_canvas.drawString(550, y, str(order.id))  # ID
            pdf_canvas.drawString(650, y, order.closuredate.strftime("%Y-%m-%d") if order.closuredate else "")  # Date
            pdf_canvas.drawString(750, y, confirmed_entry.quotation or "")  # Quotation

            y -= 20
            if y < 50:  # If reaching bottom of the page, start a new page
                pdf_canvas.showPage()
                y = 800

    pdf_canvas.save()
    return response

def export_lost_enquiries_csv(request):
    current_user = request.user
    if current_user.is_superuser:
        lost_enquiries = Enquiry.objects.filter(is_lost=True)
        relegated_enquiries = Enquiry.objects.filter(is_relegated=True)
    else:
        lost_enquiries = Enquiry.objects.filter(is_lost=True, created_by=current_user)
        relegated_enquiries = Enquiry.objects.filter(
            is_relegated=True,
            created_by=current_user
        )
    combined_enquiries = lost_enquiries.union(relegated_enquiries)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="lost_enquiries.csv"'

    writer = csv.writer(response)
    writer.writerow(['ID', 'Company Name', 'Customer Name', 'Email', 'Contact', 'Location', 'Products', 'Remarks', 'Reason for Del'])

    for enquiry in lost_enquiries:
        writer.writerow([enquiry.id, enquiry.companyname, enquiry.customername, enquiry.email, enquiry.contact, enquiry.location, enquiry.products.name, enquiry.remarks, enquiry.flag])

    return response


def export_lost_enquiries_xlsx(request):
    current_user = request.user
    if current_user.is_superuser:
        lost_enquiries = Enquiry.objects.filter(is_lost=True)
        relegated_enquiries = Enquiry.objects.filter(is_relegated=True)
    else:
        lost_enquiries = Enquiry.objects.filter(is_lost=True, created_by=current_user)
        relegated_enquiries = Enquiry.objects.filter(
            is_relegated=True,
            created_by=current_user
        )
    combined_enquiries = lost_enquiries.union(relegated_enquiries)

    wb = Workbook()
    ws = wb.active
    ws.title = "Lost Enquiries"

    # Write header row
    ws.append(['ID', 'Company Name', 'Customer Name', 'Email', 'Contact', 'Location', 'Products', 'Remarks', 'Reason for Del'])

    for enquiry in combined_enquiries:
        ws.append([enquiry.id, enquiry.companyname, enquiry.customername, enquiry.email, enquiry.contact, enquiry.location, enquiry.products.name, enquiry.remarks, enquiry.flag])

    # Create an in-memory file and save the workbook
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    # Return the response
    response = HttpResponse(stream, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="lost_enquiries.xlsx"'

    return response

def export_lost_enquiries_pdf(request):
    from io import BytesIO
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas

    current_user = request.user
    if current_user.is_superuser:
        lost_enquiries = Enquiry.objects.filter(is_lost=True)
        relegated_enquiries = Enquiry.objects.filter(is_relegated=True)
    else:
        lost_enquiries = Enquiry.objects.filter(is_lost=True, created_by=current_user)
        relegated_enquiries = Enquiry.objects.filter(
            is_relegated=True,
            created_by=current_user
        )
    combined_enquiries = lost_enquiries.union(relegated_enquiries)
    print(combined_enquiries)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="lost_enquiries.pdf"'

    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    # Define some fonts and styles
    p.setFont("Helvetica", 10)

    # Table header
    p.drawString(30, height - 40, "ID")
    p.drawString(70, height - 40, "Company Name")
    p.drawString(200, height - 40, "Customer Name")
    p.drawString(350, height - 40, "Email")
    p.drawString(500, height - 40, "Contact")
    p.drawString(600, height - 40, "Location")
    p.drawString(750, height - 40, "Products")

    y_position = height - 60
    for enquiry in combined_enquiries:
        p.drawString(30, y_position, str(enquiry.id))
        p.drawString(70, y_position, str(enquiry.companyname))
        p.drawString(200, y_position, str(enquiry.customername))
        p.drawString(350, y_position, str(enquiry.email))
        p.drawString(500, y_position, str(enquiry.contact))
        p.drawString(600, y_position, str(enquiry.location))
        p.drawString(750, y_position, str(enquiry.products.name))

        y_position -= 20

    p.showPage()
    p.save()

    # Save to response
    buffer.seek(0)
    response.write(buffer.getvalue())
    return response








# ---------------------------------------------------------------------------------------------------------------
import csv
from django.http import HttpResponse
from django.db.models import Q
from .models import Enquiry, User

def export_enquiries_csv(request):
    current_user = request.user
    selected_user_id = request.GET.get('user')
    selected_user = current_user

    # If a specific user is selected
    if selected_user_id:
        selected_user = User.objects.filter(id=selected_user_id).first()

        # Managers and superusers can see all enquiries (non-confirmed, non-lost, or reverted)
        enquiries = Enquiry.objects.filter(
            Q(is_confirmed=False, is_lost=False) | Q(is_reverted=True)
        )
    else:
        # Other users see only their own created or assigned enquiries
        enquiries = Enquiry.objects.filter(
            Q(is_confirmed=False, is_lost=False) | Q(is_reverted=True)
        ).filter(
            Q(created_by=current_user) | Q(executive__name=current_user.username)
        )
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="enquiries.csv"'

    writer = csv.writer(response)
    writer.writerow(['ID', 'Company Name', 'Customer Name', 'Contact', 'Status', 'Product', 'Enquiry Date', 'Remarks'])
    
    # Write data to CSV
    for enquiry in enquiries:
        writer.writerow([
            enquiry.id,
            enquiry.companyname,
            enquiry.customername,
            enquiry.contact,
            enquiry.get_status_display(),
            enquiry.products.name,
            enquiry.closuredate,
            enquiry.remarks
        ])
    
    return response


from io import BytesIO
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from django.db.models import Q
from .models import Enquiry, User,CommercialQuote

def export_enquiries_pdf(request):
    current_user = request.user
    selected_user_id = request.GET.get('user')
    selected_user = current_user

    if selected_user_id:
        selected_user = User.objects.filter(id=selected_user_id).first()
        # Managers/superusers see all enquiries (non-confirmed, non-lost, or reverted)
        enquiries = Enquiry.objects.filter(
            Q(is_confirmed=False, is_lost=False) | Q(is_reverted=True)
        )
    else:
        # Other users see their own created or assigned enquiries
        enquiries = Enquiry.objects.filter(
            Q(is_confirmed=False, is_lost=False) | Q(is_reverted=True)
        ).filter(
            Q(created_by=current_user) | Q(executive__name=current_user.username)
        )
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="enquiries.pdf"'

    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    # Table headers
    p.setFont("Helvetica", 10)
    p.drawString(30, height - 40, "ID")
    p.drawString(70, height - 40, "Company Name")
    p.drawString(200, height - 40, "Customer Name")
    p.drawString(350, height - 40, "Contact")
    p.drawString(450, height - 40, "Status")
    p.drawString(500, height - 40, "Product")
    p.drawString(600, height - 40, "Enquiry Date")
    p.drawString(700, height - 40, "Remarks")

    y_position = height - 60
    for enquiry in enquiries:
        p.drawString(30, y_position, str(enquiry.id))
        p.drawString(70, y_position, str(enquiry.companyname))
        p.drawString(200, y_position, str(enquiry.customername))
        p.drawString(350, y_position, str(enquiry.contact))
        p.drawString(450, y_position, str(enquiry.get_status_display()))
        p.drawString(500, y_position, str(enquiry.products.name))
        p.drawString(600, y_position, str(enquiry.closuredate))
        p.drawString(700, y_position, str(enquiry.remarks))

        y_position -= 20

    p.showPage()
    p.save()

    buffer.seek(0)
    response.write(buffer.getvalue())
    return response


import openpyxl
from django.http import HttpResponse
from django.db.models import Q
from .models import Enquiry, User

def export_enquiries_xlsx(request):
    current_user = request.user
    selected_user_id = request.GET.get('user')
    selected_user = current_user

    if selected_user_id:
        selected_user = User.objects.filter(id=selected_user_id).first()
        # Managers/superusers see all enquiries (non-confirmed, non-lost, or reverted)
        enquiries = Enquiry.objects.filter(
            Q(is_confirmed=False, is_lost=False) | Q(is_reverted=True)
        )
    else:
        # Other users see only their own created or assigned enquiries
        enquiries = Enquiry.objects.filter(
            Q(is_confirmed=False, is_lost=False) | Q(is_reverted=True)
        ).filter(
            Q(created_by=current_user) | Q(executive__name=current_user.username)
        )

    # Create in-memory workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Enquiries"
    
    # Add headers
    sheet.append(['ID', 'Company Name', 'Customer Name', 'Contact', 'Status', 'Product', 'Enquiry Date', 'Remarks'])

    for enquiry in enquiries:
        sheet.append([
            enquiry.id,
            enquiry.companyname,
            enquiry.customername,
            enquiry.contact,
            enquiry.get_status_display(),
            enquiry.products.name,
            enquiry.closuredate,
            enquiry.remarks
        ])

    # Save to response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="enquiries.xlsx"'

    workbook.save(response)
    return response

def export_quotation_data(request, enquiry_id):
    enquiry = get_object_or_404(Enquiry, id=enquiry_id)
    
    # Fetch quotations
    quotations = CommercialQuote.objects.filter(enquiry_id=enquiry_id).order_by('-id')
    
    # Initialize file paths for 'stored_data' and 'proposal'
    base_dir = settings.BASE_DIR
    stored_data_path = os.path.join(base_dir, 'stored_data')
    proposal_path = os.path.join(base_dir, 'proposal')
    draft_path = os.path.join(base_dir, 'AMC_draft')
    proposal_draft_path = os.path.join(base_dir, 'proposal_draft')

    # Data containers for each category
    stored_data = defaultdict(list)
    proposal_data = defaultdict(list)
    draft_data = defaultdict(list)
    proposal_draft_data = defaultdict(list)

    # Collect data from 'stored_data'
    stored_data_directories = os.listdir(stored_data_path)
    for directory in stored_data_directories:
        dir_path = os.path.join(stored_data_path, directory)
        if os.path.isdir(dir_path) and directory == str(enquiry_id):
            json_files = os.listdir(dir_path)
            for json_file in json_files:
                file_path = os.path.join(dir_path, json_file)
                if os.path.exists(file_path):
                    with open(file_path, 'r') as file:
                        try:
                            data = json.load(file)
                            if isinstance(data, dict) and 'enquiry_id' in data and str(data['enquiry_id']) == str(enquiry_id):
                                stored_data[os.path.basename(file_path)].append(data)
                        except json.JSONDecodeError:
                            stored_data[os.path.basename(file_path)].append({'error': 'Invalid JSON file'})

    # Collect data from 'proposal'
    proposal_data_directories = os.listdir(proposal_path)
    for directory in proposal_data_directories:
        dir_path = os.path.join(proposal_path, directory)
        if os.path.isdir(dir_path) and directory == str(enquiry_id):
            json_files = os.listdir(dir_path)
            for json_file in json_files:
                file_path = os.path.join(dir_path, json_file)
                if os.path.exists(file_path):
                    with open(file_path, 'r') as file:
                        try:
                            data = json.load(file)
                            if isinstance(data, dict) and 'enquiry_id' in data and str(data['enquiry_id']) == str(enquiry_id):
                                proposal_data[os.path.basename(file_path)].append(data)
                        except json.JSONDecodeError:
                            proposal_data[os.path.basename(file_path)].append({'error': 'Invalid JSON file'})

    # Collect data from 'AMC_draft'
    draft_data_directories = os.listdir(draft_path)
    for directory in draft_data_directories:
        dir_path = os.path.join(draft_path, directory)
        if os.path.isdir(dir_path) and directory == str(enquiry_id):
            json_files = os.listdir(dir_path)
            for json_file in json_files:
                file_path = os.path.join(dir_path, json_file)
                if os.path.exists(file_path):
                    with open(file_path, 'r') as file:
                        try:
                            data = json.load(file)
                            if isinstance(data, dict) and 'enquiry_id' in data and str(data['enquiry_id']) == str(enquiry_id):
                                draft_data[os.path.basename(file_path)].append(data)
                        except json.JSONDecodeError:
                            draft_data[os.path.basename(file_path)].append({'error': 'Invalid JSON file'})

    # Collect data from 'proposal_draft'
    proposal_draft_data_directories = os.listdir(proposal_draft_path)
    for directory in proposal_draft_data_directories:
        dir_path = os.path.join(proposal_draft_path, directory)
        if os.path.isdir(dir_path) and directory == str(enquiry_id):
            json_files = os.listdir(dir_path)
            for json_file in json_files:
                file_path = os.path.join(dir_path, json_file)
                if os.path.exists(file_path):
                    with open(file_path, 'r') as file:
                        try:
                            data = json.load(file)
                            if isinstance(data, dict) and 'enquiry_id' in data and str(data['enquiry_id']) == str(enquiry_id):
                                proposal_draft_data[os.path.basename(file_path)].append(data)
                        except json.JSONDecodeError:
                            proposal_draft_data[os.path.basename(file_path)].append({'error': 'Invalid JSON file'})

    # Prepare CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="quotation_numbers_{enquiry_id}.csv"'
    writer = csv.writer(response)

    # Write the header row
    writer.writerow(['Category', 'Quotation Number'])

    # Write data for 'stored_data'
    for data_list in stored_data.values():
        for data in data_list:
            quotation_number = data.get('quotation_number', 'N/A')
            writer.writerow(['AMC ', quotation_number])

    # Write data for 'proposal_data'
    for data_list in proposal_data.values():
        for data in data_list:
            quotation_number = data.get('quotation_number', 'N/A')
            writer.writerow(['Proposal ', quotation_number])

    # Write data for 'draft_data'
    for data_list in draft_data.values():
        for data in data_list:
            quotation_number = data.get('quotation_number', 'N/A')
            writer.writerow([' AMC Draft ', quotation_number])

    # Write data for 'proposal_draft_data'
    for data_list in proposal_draft_data.values():
        for data in data_list:
            quotation_number = data.get('quotation_number', 'N/A')
            writer.writerow(['Proposal Draft Data', quotation_number])

    for quote in quotations:
        writer.writerow(['Commercial Quote', quote.quotation_no])

    return response




from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from django.http import HttpResponse
from collections import defaultdict
import os
import json
from django.shortcuts import get_object_or_404
from .models import Enquiry, CommercialQuote
from django.conf import settings

def export_quotation_pdf(request, enquiry_id):
    enquiry = get_object_or_404(Enquiry, id=enquiry_id)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="quotation_numbers_{enquiry_id}.pdf"'

    # Set up the PDF document
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []

    # Data containers for different categories
    stored_data_path = os.path.join(settings.BASE_DIR, 'stored_data')
    proposal_path = os.path.join(settings.BASE_DIR, 'proposal')
    draft_path = os.path.join(settings.BASE_DIR, 'AMC_draft')
    proposal_draft_path = os.path.join(settings.BASE_DIR, 'proposal_draft')

    stored_data = defaultdict(list)
    proposal_data = defaultdict(list)
    draft_data = defaultdict(list)
    proposal_draft_data = defaultdict(list)

    # Function to collect data from directories
    def collect_data(path, category, data_container):
        directories = os.listdir(path)
        for directory in directories:
            dir_path = os.path.join(path, directory)
            if os.path.isdir(dir_path) and directory == str(enquiry_id):
                json_files = os.listdir(dir_path)
                for json_file in json_files:
                    file_path = os.path.join(dir_path, json_file)
                    if os.path.exists(file_path):
                        with open(file_path, 'r') as file:
                            try:
                                data = json.load(file)
                                if isinstance(data, dict) and 'enquiry_id' in data and str(data['enquiry_id']) == str(enquiry_id):
                                    # Now append to the list under the dictionary key
                                    data_container[category].append(data)
                            except json.JSONDecodeError:
                                data_container[category].append({'error': 'Invalid JSON file'})

    # Collect all data
    collect_data(stored_data_path, 'AMC', stored_data)
    collect_data(proposal_path, 'Proposal', proposal_data)
    collect_data(draft_path, 'AMC Draft', draft_data)
    collect_data(proposal_draft_path, 'Proposal Draft', proposal_draft_data)

    # Prepare table data
    data = [["Category", "Quotation Numbers"]]

    # Adding rows for each category and its respective quotation numbers
    def add_category_data(category, data_container):
        for data_list in data_container.get(category, []):
            quotation_number = data_list.get('quotation_number', 'N/A')
            data.append([category, quotation_number])

    add_category_data("AMC", stored_data)
    add_category_data("Proposal", proposal_data)
    add_category_data("AMC Draft", draft_data)
    add_category_data("Proposal Draft", proposal_draft_data)

    # Adding commercial quotes to the table
    quotations = CommercialQuote.objects.filter(enquiry_id=enquiry_id).order_by('-id')
    for quote in quotations:
        data.append(["Commercial Quote", quote.quotation_no])

    # Create the table
    table = Table(data)

    # Set table style (e.g., borders, alignment)
    table.setStyle(TableStyle([
        ('TEXTCOLOR', (0, 0), (-1, 0), (0, 0, 0)),  # Table header color
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Align all columns center
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Bold header
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),  # Padding for header row
        ('BACKGROUND', (0, 0), (-1, 0), (0.8, 0.8, 0.8)),  # Background color for header
        ('GRID', (0, 0), (-1, -1), 1, (0, 0, 0)),  # Add gridlines
    ]))

    # Append table to the document elements
    elements.append(table)

    # Build the PDF document
    doc.build(elements)

    return response




from openpyxl import Workbook
from django.http import HttpResponse

def export_quotation_excel(request, enquiry_id):
    enquiry = get_object_or_404(Enquiry, id=enquiry_id)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="quotation_numbers_{enquiry_id}.xlsx"'

    # Create a new workbook and active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotations"

    # Write headers
    ws.append(['Category', 'Quotation Number'])

    # Fetch quotations and organize data as in the CSV and PDF views
    base_dir = settings.BASE_DIR
    stored_data_path = os.path.join(base_dir, 'stored_data')
    proposal_path = os.path.join(base_dir, 'proposal')
    draft_path = os.path.join(base_dir, 'AMC_draft')
    proposal_draft_path = os.path.join(base_dir, 'proposal_draft')

    stored_data = defaultdict(list)
    proposal_data = defaultdict(list)
    draft_data = defaultdict(list)
    proposal_draft_data = defaultdict(list)

    # Collect data (use the same logic as above)
    def collect_data(path, category, data_container):
        directories = os.listdir(path)
        for directory in directories:
            dir_path = os.path.join(path, directory)
            if os.path.isdir(dir_path) and directory == str(enquiry_id):
                json_files = os.listdir(dir_path)
                for json_file in json_files:
                    file_path = os.path.join(dir_path, json_file)
                    if os.path.exists(file_path):
                        with open(file_path, 'r') as file:
                            try:
                                data = json.load(file)
                                if isinstance(data, dict) and 'enquiry_id' in data and str(data['enquiry_id']) == str(enquiry_id):
                                    data_container[os.path.basename(file_path)].append(data)
                            except json.JSONDecodeError:
                                data_container[os.path.basename(file_path)].append({'error': 'Invalid JSON file'})

    collect_data(stored_data_path, 'AMC', stored_data)
    collect_data(proposal_path, 'Proposal', proposal_data)
    collect_data(draft_path, 'AMC Draft', draft_data)
    collect_data(proposal_draft_path, 'Proposal Draft', proposal_draft_data)

    # Write data into Excel
    for data_list in stored_data.values():
        for data in data_list:
            quotation_number = data.get('quotation_number', 'N/A')
            ws.append(['AMC', quotation_number])

    for data_list in proposal_data.values():
        for data in data_list:
            quotation_number = data.get('quotation_number', 'N/A')
            ws.append(['Proposal', quotation_number])

    for data_list in draft_data.values():
        for data in data_list:
            quotation_number = data.get('quotation_number', 'N/A')
            ws.append(['AMC Draft', quotation_number])

    for data_list in proposal_draft_data.values():
        for data in data_list:
            quotation_number = data.get('quotation_number', 'N/A')
            ws.append(['Proposal Draft', quotation_number])

    quotations = CommercialQuote.objects.filter(enquiry_id=enquiry_id).order_by('-id')
    for quote in quotations:
        ws.append(['Commercial Quote', quote.quotation_no])

    # Save the workbook to the response
    wb.save(response)
    return response
