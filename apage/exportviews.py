# exportview.py
from django.http import HttpResponse
from .models import GeneralReport
import csv
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

def export_data(request):
    # Get the report type and format from the request
    record_type = request.GET.get('type', 'general_report')
    file_format = request.GET.get('format', 'csv')

    if record_type == 'general_report':
        queryset = GeneralReport.objects.all()
        fields = ['Site', 'Date of Visit', 'Point 1', 'Point 2', 'Point 3', 'Point 4', 'Notes', 'Attachment']
    else:
        return HttpResponse("Unsupported record type", status=400)

    # Call the appropriate export function based on the format requested
    if file_format == 'csv':
        return export_csv(request, queryset, fields)
    elif file_format == 'xlsx':
        return export_xlsx(request, queryset, fields)
    elif file_format == 'pdf':
        return export_pdf(request, queryset, fields)
    else:
        return HttpResponse("Unsupported format", status=400)

def export_csv(request, queryset, fields):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="general_report_export.csv"'
    writer = csv.writer(response)
    
    # Write header row
    writer.writerow(fields)
    
    # Write the data for each report
    for report in queryset:
        writer.writerow([
            report.site.name, 
            report.date_of_visit, 
            report.point1, 
            report.point2, 
            report.point3, 
            report.point4, 
            report.notes, 
            report.attachment.url if report.attachment else 'No Attachment'
        ])
    return response

def export_xlsx(request, queryset, fields):
    wb = Workbook()
    ws = wb.active
    ws.title = "General Report Export"
    
    # Write header row
    ws.append(fields)

    # Write data rows
    for report in queryset:
        ws.append([
            report.site.name, 
            report.date_of_visit, 
            report.point1, 
            report.point2, 
            report.point3, 
            report.point4, 
            report.notes, 
            report.attachment.url if report.attachment else 'No Attachment'
        ])
    
    # Prepare response with XLSX content type
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="general_report_export.xlsx"'
    
    wb.save(response)
    return response

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle
from django.http import HttpResponse

def export_pdf(request, queryset, fields):
    # Create an HTTP response with PDF content
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="general_report_export.pdf"'
    
    # Create a PDF canvas
    c = canvas.Canvas(response, pagesize=letter)
    
    # Define the table data (including header and rows)
    table_data = []
    
    # Add the header row
    table_data.append(fields)

    # Add data rows
    for report in queryset:
        table_data.append([
            report.site.name,
            report.date_of_visit,
            report.point1,
            report.point2,
            report.point3,
            report.point4,
            report.notes,
            'Yes' if report.attachment else 'No Attachment'
        ])
    
    # Create a Table object and apply some styles
    table = Table(table_data, colWidths=[150, 80, 100, 100, 100, 100, 150, 100])  # Adjust column widths as needed
    table.setStyle(TableStyle([
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),  # Header text color
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Center align all cells
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Header font style
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),  # Padding for header row
        ('BACKGROUND', (0, 0), (-1, 0), colors.black),  # Header background color
        ('GRID', (0, 0), (-1, -1), 1, colors.black),  # Grid lines for the table
        ('BOX', (0, 0), (-1, -1), 0.5, colors.black),  # Box around the table
    ]))
    
    # Build the table in the PDF document
    table.wrapOn(c, 50, 500)
    table.drawOn(c, 50, 500)
    
    c.showPage()  # Start a new page in the PDF
    c.save()  # Save the PDF document

    return response
# ---------------------------------mom-----------------------------------

from django.http import HttpResponse
from django.shortcuts import render
import csv
import xlsxwriter
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
from .models import MOM

def export_mom(request):
    format_type = request.GET.get('format')
    mom_records = MOM.objects.all()

    if format_type == 'csv':
        return export_mom_csv(mom_records)
    elif format_type == 'xlsx':
        return export_mom_xlsx(mom_records)
    elif format_type == 'pdf':
        return export_mom_pdf(mom_records)

def export_mom_csv(mom_records):
    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="mom_records.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Topic', 'Organized By', 'Date', 'Location', 'Meeting Chair', 'Attachments'])
    
    for mom in mom_records:
        writer.writerow([mom.topic, mom.organize, mom.date, mom.location, mom.meeting_chair, ', '.join([attachment.file.name for attachment in mom.attachments.all()])])
    
    return response

def export_mom_xlsx(mom_records):
    # Create XLSX response
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet()

    # Set the column headers
    worksheet.write(0, 0, 'Topic')
    worksheet.write(0, 1, 'Organized By')
    worksheet.write(0, 2, 'Date')
    worksheet.write(0, 3, 'Location')
    worksheet.write(0, 4, 'Meeting Chair')
    worksheet.write(0, 5, 'Attachments')

    row = 1
    for mom in mom_records:
        worksheet.write(row, 0, mom.topic)
        worksheet.write(row, 1, mom.organize)
        worksheet.write(row, 2, str(mom.date))
        worksheet.write(row, 3, mom.location)
        worksheet.write(row, 4, mom.meeting_chair)
        worksheet.write(row, 5, ', '.join([attachment.file.name for attachment in mom.attachments.all()]))
        row += 1

    workbook.close()
    output.seek(0)

    response = HttpResponse(output.read(), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="mom_records.xlsx"'
    return response

def export_mom_pdf(mom_records):
    # Create PDF response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="mom_records.pdf"'
    
    c = canvas.Canvas(response, pagesize=letter)
    data = [['Topic', 'Organized By', 'Date', 'Location', 'Meeting Chair', 'Attachments']]
    
    for mom in mom_records:
        data.append([mom.topic, mom.organize, mom.date, mom.location, mom.meeting_chair, ', '.join([attachment.file.name for attachment in mom.attachments.all()])])

    table = Table(data)
    table.setStyle(TableStyle([
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke), 
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.blue),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))

    table.wrapOn(c, 50, 500)
    table.drawOn(c, 50, 500)
    
    c.showPage()
    c.save()

    return response
# -----------------------------maintenancechecklist--------------------
import csv
from django.http import HttpResponse
from .models import MaintenanceChecklist

def export_maintenance_checklist_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="maintenance_checklist_records.csv"'

    writer = csv.writer(response)
    writer.writerow(['Inspector Name', 'Date', 'Machine', 'Visit Date', 'Supply Voltage', 'Current Load', 'Observations'])

    checklists = MaintenanceChecklist.objects.all()
    for checklist in checklists:
        writer.writerow([
            checklist.inspector_name,
            checklist.date,
            checklist.machine.name,
            checklist.visit_date,
            checklist.supply_voltage,
            checklist.current_load,
            checklist.observations
        ])

    return response
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from .models import MaintenanceChecklist
from django.http import HttpResponse
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from .models import MaintenanceChecklist

def export_maintenance_checklist_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="maintenance_checklist_records.pdf"'

    # Create PDF document
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []  # Holds the elements of the PDF

    # Add Title
    styles = getSampleStyleSheet()
    title = Paragraph("Maintenance Checklist Records", styles['Title'])
    elements.append(title)

    # Table Header
    data = [
        ['Inspector Name', 'Date', 'Machine', 'Visit Date', 'Observations']
    ]

    # Fetch Data
    checklists = MaintenanceChecklist.objects.all()
    for checklist in checklists:
        row = [
            checklist.inspector_name,
            checklist.date.strftime('%Y-%m-%d') if checklist.date else '',
            checklist.machine.name,
            checklist.visit_date.strftime('%Y-%m-%d') if checklist.visit_date else '',
            checklist.observations,
        ]
        data.append(row)

    # Create Table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),  # Header background color
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),  # Header text color
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Center align all cells
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Bold header font
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),  # Add padding for header
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),  # Alternate row color
        ('GRID', (0, 0), (-1, -1), 1, colors.black),  # Grid lines
    ]))

    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    return response

import openpyxl
from django.http import HttpResponse
from .models import MaintenanceChecklist

def export_maintenance_checklist_xlsx(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="maintenance_checklist_records.xlsx"'

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Maintenance Checklist Records"

    # Header row
    headers = ['Inspector Name', 'Date', 'Machine', 'Visit Date', 'Supply Voltage', 'Current Load']
    sheet.append(headers)

    # Data rows
    checklists = MaintenanceChecklist.objects.all()
    for checklist in checklists:
        row = [
            checklist.inspector_name,
            checklist.date,
            checklist.machine.name,
            checklist.visit_date,
            checklist.supply_voltage,
            checklist.current_load
        ]
        sheet.append(row)

    workbook.save(response)
    return response

# ----------------------------servicereport---------------------------------------
from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from .models import ServiceReport

def export_service_reports_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="service_reports.pdf"'

    # Initialize PDF document
    doc = SimpleDocTemplate(response, pagesize=landscape(letter))
    elements = []

    # Data for table header and rows
    data = [['Service Name', 'Date of Visit', 'Zone', 'Phone No', 'Reason of Visit', 'In Time', 'Out Time', 'Certified By']]

    # Fetch all service reports
    reports = ServiceReport.objects.filter(service_name=request.user.username)
    for report in reports:
        data.append([
            report.service_name,
            report.date_of_visit.strftime('%Y-%m-%d'),
            report.get_zone_display(),
            report.phone_no,
            report.reason_of_visit,
            report.in_time.strftime('%H:%M:%S'),
            report.out_time.strftime('%H:%M:%S'),
            report.certified_by_name or '',
        ])

    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
    ]))

    elements.append(table)
    doc.build(elements)

    return response
import csv
from django.http import HttpResponse
from .models import ServiceReport

def export_service_reports_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="service_reports.csv"'

    writer = csv.writer(response)
    # Write table headers
    writer.writerow(['Service Name', 'Date of Visit', 'Zone', 'Phone No', 'Reason of Visit', 'In Time', 'Out Time', 'Certified By'])

    # Write data rows
    reports = ServiceReport.objects.all()
    for report in reports:
        writer.writerow([
            report.service_name,
            report.date_of_visit,
            report.get_zone_display(),
            report.phone_no,
            report.reason_of_visit,
            report.in_time,
            report.out_time,
            report.certified_by_name or '',
        ])

    return response
from django.http import HttpResponse
from openpyxl import Workbook
from .models import ServiceReport

def export_service_reports_xlsx(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="service_reports.xlsx"'

    # Initialize workbook
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Service Reports'

    # Write table headers
    headers = ['Service Name', 'Date of Visit', 'Zone', 'Phone No', 'Reason of Visit', 'In Time', 'Out Time', 'Certified By']
    worksheet.append(headers)

    # Write data rows
    reports = ServiceReport.objects.all()
    for report in reports:
        worksheet.append([
            report.service_name,
            report.date_of_visit.strftime('%Y-%m-%d'),
            report.get_zone_display(),
            report.phone_no,
            report.reason_of_visit,
            report.in_time.strftime('%H:%M:%S'),
            report.out_time.strftime('%H:%M:%S'),
            report.certified_by_name or '',
        ])

    # Save workbook to response
    workbook.save(response)
    return response

