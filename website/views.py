from collections import defaultdict
import decimal
from importlib.metadata import files
import os
import uuid
from django.conf import settings
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.core.exceptions import ValidationError
from requests import request
from website.fcm_utils import get_access_token
from website.models import AdvanceAssignment, AdvanceGroup, BorrowedAmount, Conveyance, Expense, Notification, PaymentRequestMessage, ProofPhoto, AdvanceGroupUpdateLog
from django.core.files.storage import FileSystemStorage  # Added for file storage handling
from django.http import HttpResponse
from django.shortcuts import redirect
from django.shortcuts import render, redirect
from .forms import ExpenseForm
from pytesseract import Output

from django.views.decorators.http import require_POST
from django.http import JsonResponse
from .models import CashVoucher, FCMToken

from django.views.decorators.http import require_POST
from django.http import JsonResponse
from .models import CashVoucher, Expense


from django.db.models import Q




@require_POST
def approve_voucher_ajax(request):
    voucher_id = request.POST.get('voucher_id')
    try:
        voucher = CashVoucher.objects.get(id=voucher_id)
        voucher.status = 'approved'
        voucher.save()

        # ✅ Update the related expense amount
        if voucher.expense and voucher.amount:
            expense = voucher.expense
            expense.amount = (expense.amount or 0) + voucher.amount
            expense.save()

        return JsonResponse({'success': True})
    except CashVoucher.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Voucher not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

def success_view(request):
    return render(request, 'xp/success.html')  # Success page after form submission

from .models import Notification
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from datetime import timedelta
from django.utils import timezone
from apage.models import ServiceReport, Site


@login_required(login_url='/login/')
def home(request):
    today = timezone.now().date()

    all_sites = Site.objects.all()
    visit_data = []

    # Existing logic you already have...
    for site in all_sites:
        last_visit = (
            ServiceReport.objects.filter(site=site)
            .order_by('-date_of_visit')
            .first()
        )

        if last_visit:
            next_visit = last_visit.date_of_visit + timedelta(days=15)
            is_due = next_visit < today
            visit_data.append({
                'site': site,
                'last_visit': last_visit.date_of_visit,
                'next_visit': next_visit,
                'is_due': is_due,
            })

    # Notifications
    if request.user.is_staff or request.user.is_superuser:
        notifications = Notification.objects.all().order_by('-created_at')
    else:
        notifications = Notification.objects.filter(user=request.user).order_by('-created_at')

    # -------------------------------
    # 🎉 BIRTHDAY LOGIC
    # -------------------------------

    from django.contrib.auth.models import User

    # Users whose birthday is today
    todays_birthdays = User.objects.filter(
        userprofile__date_of_birth__month=today.month,
        userprofile__date_of_birth__day=today.day
    )

    # Current user's birthday?
    self_birthday = todays_birthdays.filter(id=request.user.id).exists()

    # Others birthday (exclude logged-in user)
    others_birthdays = todays_birthdays.exclude(id=request.user.id)

    # 🎈 Show popup only once per login
    show_birthday_popup = False
    if (self_birthday or others_birthdays.exists()) and not request.session.get("birthday_popup_shown"):
        show_birthday_popup = True
        request.session["birthday_popup_shown"] = True

    return render(request, 'xp/home.html', {
        'notifications': notifications,
        'visit_data': visit_data,
        'todays_birthdays': todays_birthdays,
        'self_birthday': self_birthday,
        'others_birthdays': others_birthdays,
        'show_birthday_popup': show_birthday_popup,
    })

def base(request):

    return render(request, 'xp/base.html')

from django.shortcuts import render, redirect, get_object_or_404
from .models import Expense, CashVoucher, Conveyance
from django.contrib.auth.models import User
from django.contrib import messages
from django.http import HttpResponse
from datetime import datetime
from django.db.models import Q, Count
import logging
from django.contrib.auth.decorators import login_required
from bs4 import BeautifulSoup
@login_required


def item_form(request):
    print("Entered item_form view") 
    user = request.user
    search_query = request.GET.get('q', '').strip()
    transaction_date_from = request.GET.get('transaction_date_from', None)
    transaction_date_to = request.GET.get('transaction_date_to', None)
    item_type_filter = request.GET.get('item_type', '')
    sort_by_item_type = request.GET.get('sort_by_item_type', '0').strip()
    sort_delete_requested = request.GET.get('sort_delete_requested', '0').strip()
    sort_approval_status = request.GET.get('sort_approval_status', '0').strip()
    sort_payment_status = request.GET.get('sort_payment_status', '0').strip()
    logger = logging.getLogger(__name__)
    selected_user_id = request.GET.get('user_id', None)
    page_number = request.GET.get('page') or 1

    logger.debug(f"User: {user.username}, is_staff: {user.is_staff}, is_superuser: {user.is_superuser}")
    logger.debug(f"Search Query: {search_query}")
    logger.debug(f"Date Range: From {transaction_date_from} to {transaction_date_to}")

    # Base queryset
    if user.is_staff or user.is_superuser:
        print("ADMIN LOGIN FILTER")
        expenses = Expense.objects.all()
        all_users = User.objects.all()
        if selected_user_id:
            expenses = expenses.filter(created_by__id=selected_user_id)
            try:
                selected_user = User.objects.get(id=selected_user_id)
            except User.DoesNotExist:
                selected_user = None
        else:
            selected_user = None
    else:
        print("NON ADMIN LOGIN")
        expenses = Expense.objects.filter(
            Q(created_by=user) |
            Q(transaction_category='internal', internal_option=user.username) |
            Q(transaction_category='internal', internal_option=user.first_name)
        )
        all_users = None
        selected_user = None
        logger.debug("User viewing their own and assigned internal expenses")

    # Apply item type filter
    if item_type_filter:
        print("ITEM TYPE FILTER APPLIED")
        expenses = expenses.filter(item_type=item_type_filter)

    # Annotate proof count
    expenses = expenses.annotate(proof_count=Count('proof_photos'))
    # Search query
    if search_query:
        print("search query applied")
        expenses = expenses.filter(
            Q(item_type__icontains=search_query) |
            Q(item_name__icontains=search_query) |
            Q(transaction_category__icontains=search_query) |
            Q(external_type__icontains=search_query) |
            Q(internal_option__icontains=search_query) |
            Q(payment_mode__icontains=search_query) |
            Q(amount__icontains=search_query) |
            Q(transaction_date__icontains=search_query) |
            Q(created_by__username__icontains=search_query) |
            Q(amount__icontains=search_query) |
            Q(evoucher_number__icontains=search_query)
        )


    # Apply date filters
    if transaction_date_from:
        print("trasaction date from filter applied")
        try:
            transaction_date_from = datetime.strptime(transaction_date_from, '%Y-%m-%d').date()
            expenses = expenses.filter(transaction_date__gte=transaction_date_from)
        except ValueError:
            logger.warning("Invalid format for 'transaction_date_from'")
    if transaction_date_to:
        print("trasaction date to filter applied")
        try:
            transaction_date_to = datetime.strptime(transaction_date_to, '%Y-%m-%d').date()
            expenses = expenses.filter(transaction_date__lte=transaction_date_to)
        except ValueError:
            logger.warning("Invalid format for 'transaction_date_to'")

        # --- Sorting Logic ---
    print(f"SORT FLAGS → approval={sort_approval_status}, delete={sort_delete_requested}, payment={sort_payment_status}")

    if sort_delete_requested == '1':
        print("🟡 Sort: delete requested (DESC)")
        expenses = expenses.filter(delete_requested=True).order_by('-id')

    elif sort_by_item_type == '1':
        print("🟡 Sort: item type (ASC)")
        expenses = expenses.order_by('item_type', '-id')

    elif sort_approval_status == '1':
        print("🟡 Sort: ONLY vouchers that need approval/rejection (no proofs)")
        expenses = expenses.filter(
            Q(is_approved=False),
            Q(is_rejected=False),
        ).filter(
            Q(proof_photo__isnull=True) | Q(proof_photo=''),
            Q(bill_photo__isnull=True) | Q(bill_photo=''),
            Q(gst_photo__isnull=True) | Q(gst_photo=''),
        ).distinct().order_by('-id')

    elif sort_payment_status == '1':
        print("🟡 Sort: payment status (ASC)")
        expenses = expenses.order_by('status', '-id')

    else:
        print("⚪ Default order applied (latest first)")
        expenses = expenses.order_by('-id')

    # Only show non-drafts
    expenses = expenses.filter(is_draft=False)
    # if request.GET.get('sort_approval_status') is not None:
    #     if sort_approval_status == '1':
    #         print("non draft filter applied")
    #         expenses = sorted(
    #             list(expenses),
    #             key=lambda e: not (
    #                 not e.is_approved
    #                 and not e.is_rejected
    #                 and ((not e.proof_photo and getattr(e, 'proof_count', 0) == 0) or (not e.bill_photo and not e.gst_photo))
    #             )
    #         )
    #     elif sort_approval_status == '0':
    #         print("draft filter applied")
    #         expenses = sorted(
    #             list(expenses),
    #             key=lambda e: (
    #                 not e.is_approved
    #                 and not e.is_rejected
    #                 and ((not e.proof_photo and getattr(e, 'proof_count', 0) == 0) or (not e.bill_photo and not e.gst_photo))
    #             )
    #         )

    if not any([sort_delete_requested, sort_by_item_type, sort_approval_status, sort_payment_status]):
        expenses = expenses.order_by('-id')

    # Approved vouchers and vehicle types
    approved_vouchers = CashVoucher.objects.filter(status='approved')
    conveyance_vehicle_types = [option.vehicle_type for option in Conveyance.objects.all()]

    # Add pending tip flag and derive KM for conveyance items

    # ✅ Fetch default conveyance rates once (so we don't query inside the loop)
    two_wheeler = Conveyance.objects.filter(vehicle_type='2_wheeler').first()
    four_wheeler = Conveyance.objects.filter(vehicle_type='4_wheeler').first()

    two_rate = two_wheeler.price_per_km if two_wheeler else 6  # fallback if not in DB
    four_rate = four_wheeler.price_per_km if four_wheeler else 10  # fallback if not in DB

    for exp in expenses:
        exp.has_pending_tip = exp.cash_vouchers.filter(
            voucher_number__startswith="XPCV", status="pending"
        ).exists()

        # ✅ Derive KM if missing but it's conveyance
        if exp.item_type == 'conveyance' and not exp.km and exp.amount:
            try:
                # Priority 1: use linked conveyance object
                if exp.conveyance and exp.conveyance.price_per_km:
                    rate = exp.conveyance.price_per_km
                # Priority 2: use default conveyance rates from DB
                elif exp.transaction_option == '2_wheeler':
                    rate = two_rate
                elif exp.transaction_option == '4_wheeler':
                    rate = four_rate
                else:
                    rate = None

                if rate:
                    exp.derived_km = round(exp.amount / rate, 2)
                else:
                    exp.derived_km = None

            except (ZeroDivisionError, AttributeError, TypeError):
                exp.derived_km = None
        else:
            exp.derived_km = exp.km

    paginator = Paginator(expenses, 50)
    page_obj = paginator.get_page(page_number)

    # --- Generate evoucher number ---
    evoucher_number = generate_evoucher_number()
    all_users = User.objects.all().order_by('username')

    # --- Context (use page_obj.object_list, not full queryset) ---
    context = {
        'evoucher_number': evoucher_number,
        'current_date': datetime.today().date(),
        'expenses': page_obj,              # ✅ use only paginated subset
        'page_obj': page_obj,              # ✅ template pagination
        'approved_vouchers': approved_vouchers,
        'transaction_date_from': transaction_date_from,
        'transaction_date_to': transaction_date_to,
        'item_types': Expense.objects.values_list('item_type', flat=True).distinct(),
        'sort_by_item_type': sort_by_item_type,
        'sort_delete_requested': sort_delete_requested,
        'day_ranges': [15, 30, 45, 90],
        'all_users': all_users,
        'selected_user_id': selected_user_id,
        'selected_user': selected_user,
        'sort_approval_status': sort_approval_status,
        'sort_payment_status': sort_payment_status,
    }
    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        # Render only the expense section, but keep pagination
        html = render_to_string("xp/item_form.html", context, request=request)

        soup = BeautifulSoup(html, "html.parser")
        container = soup.select_one("#expenseTableContainer")

        # Get the full section (table + pagination)
        table_html = str(container) if container else ""
        return JsonResponse({"html": table_html})

    return render(request, 'xp/item_form.html', context)

def transaction(request):
    user = request.user
    # Filters
    transaction_date_from = request.GET.get('transaction_date_from', None)
    transaction_date_to = request.GET.get('transaction_date_to', None)
    item_type_filter = request.GET.get('item_type', '')
    search_query = request.GET.get('search', '').strip()
    selected_user_id = request.GET.get('user_id', None)  # Capture user selection
    users = User.objects.all()

    # Base queryset
    if user.is_staff or user.is_superuser:
        expenses = Expense.objects.all()  # Admin or staff see all vouchers
    else:
        # Include vouchers created by the user and internal vouchers assigned to the user
        expenses = Expense.objects.filter(
    created_by=user
) | Expense.objects.filter(
    transaction_category='internal', internal_option__icontains=user.username
)
    expenses = expenses.order_by('-transaction_date')

    # Filter by selected user (if any)
    if selected_user_id:
        expenses = expenses.filter(created_by_id=selected_user_id)  # Filter by selected user

    # Apply other filters
    if item_type_filter:
        expenses = expenses.filter(item_type=item_type_filter)

    if transaction_date_from:
        try:
            transaction_date_from = datetime.strptime(transaction_date_from, '%Y-%m-%d').date()
            expenses = expenses.filter(transaction_date__gte=transaction_date_from)
        except ValueError:
            pass
    if transaction_date_to:
        try:
            transaction_date_to = datetime.strptime(transaction_date_to, '%Y-%m-%d').date()
            expenses = expenses.filter(transaction_date__lte=transaction_date_to)
        except ValueError:
            pass

    # Apply search query
    if search_query:
        expenses = expenses.filter(
            Q(evoucher_number__icontains=search_query) |
            Q(item_type__icontains=search_query) |
            Q(transaction_date__icontains=search_query) |
            Q(transaction_details__icontains=search_query)
        )

    # Prepare data for the summary
    summary_data = []
    user_balances = defaultdict(int)

    for index, expense in enumerate(expenses.distinct()):
        evoucher_number = expense.evoucher_number if expense.evoucher_number else ''
        
        # Format internal users with underscores
        if expense.transaction_category == 'internal' and expense.internal_option:
            internal_users = "_".join(expense.internal_option.split(","))
            internal_or_external = f"int_{internal_users}"
        else:
            internal_or_external = "ext"

        summary = f"{evoucher_number}_{internal_or_external}"
        debit = expense.amount
        credit = expense.amount if expense.status == 'paid' else 0
        user_balances[expense.created_by] += credit - debit
        current_balance = user_balances[expense.created_by]

        summary_data.append({
            'slno': index + 1,
            'transaction_date': expense.transaction_date,
            'summary': summary,
            'debit': debit,
            'credit': credit if credit else "",
            'balance': current_balance,
            'created_by': expense.created_by,

        })

    # Handle Export Requests
    export_type = request.GET.get('export', '').lower()
    if export_type == 'csv':
        return export_csv(summary_data)
    elif export_type == 'xlsx':
        return export_xlsx(summary_data)
    elif export_type == 'pdf':
        return export_pdf(summary_data)

    # Context for rendering the template
    context = {
        'summary_data': summary_data,
        'transaction_date_from': transaction_date_from,
        'transaction_date_to': transaction_date_to,
        'item_types': Expense.objects.values_list('item_type', flat=True).distinct(),
        'search_query': search_query,
        'users': users,
        'selected_user_id': selected_user_id,  # Pass the selected user ID to the template
        'request': request 
    }

    return render(request, 'xp/voucher_summary.html', context)

import csv
from django.http import HttpResponse

def export_csv(summary_data):
    # Create a response object with CSV content type
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="voucher_summary.csv"'

    writer = csv.writer(response)
    # Write header row
    writer.writerow(['SLNO', 'User', 'Transaction Date', 'Summary', 'Debit', 'Credit', 'Balance'])

    # Write data rows
    for item in summary_data:
        writer.writerow([
            item['slno'],
            item['created_by'].username,
            item['transaction_date'],
            item['summary'],
            item['debit'],
            item['credit'],
            item['balance']
        ])
    
    return response

from openpyxl import Workbook
from django.http import HttpResponse

def export_xlsx(summary_data):
    # Create a response object with Excel content type
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="voucher_summary.xlsx"'

    # Create a workbook and a worksheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Voucher Summary'

    # Write header row
    headers = ['SLNO', 'User', 'Transaction Date', 'Summary', 'Debit', 'Credit', 'Balance']
    worksheet.append(headers)

    # Write data rows
    for item in summary_data:
        worksheet.append([
            item['slno'],
            item['created_by'].username,
            item['transaction_date'],
            item['summary'],
            item['debit'],
            item['credit'],
            item['balance']
        ])

    # Save workbook to the response object
    workbook.save(response)
    
    return response

from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from django.http import HttpResponse

def export_pdf(summary_data):
    # Create a file-like buffer to receive PDF data
    buffer = BytesIO()
    
    # Create a canvas object to generate the PDF
    c = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    # Set font
    c.setFont("Helvetica", 10)

    # Write headers
    headers = ['SLNO', 'User', 'Transaction Date', 'Summary', 'Debit', 'Credit', 'Balance']
    x_offset = 50
    y_offset = height - 40

    for i, header in enumerate(headers):
        c.drawString(x_offset + i * 90, y_offset, header)
    
    y_offset -= 20  # Move to the next line for data rows

    # Write data rows
    for item in summary_data:
        c.drawString(x_offset, y_offset, str(item['slno']))
        c.drawString(x_offset + 90, y_offset, item['created_by'].username)
        c.drawString(x_offset + 180, y_offset, str(item['transaction_date']))
        c.drawString(x_offset + 270, y_offset, item['summary'])
        c.drawString(x_offset + 360, y_offset, str(item['debit']))
        c.drawString(x_offset + 450, y_offset, str(item['credit']))
        c.drawString(x_offset + 540, y_offset, str(item['balance']))
        y_offset -= 20

    # Save the PDF
    c.showPage()
    c.save()

    # Return the generated PDF as a response
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="voucher_summary.pdf"'

    return response
def update_item(request):
    if request.method == 'POST':
        item_id = request.POST.get('id')
        item = Expense.objects.get(id=item_id)

        # Retrieve all the form data
        item.item_type = request.POST.get('item_type')
        item.item_name = request.POST.get('item_name')
        item.transaction_option = request.POST.get('transaction_option')
        item.transaction_category = request.POST.get('transaction_category')
        item.internal_option = request.POST.get('internal_option') if item.transaction_category == 'internal' else None
        item.external_type = request.POST.get('external_type') if item.transaction_category == 'external' else None
        item.amount = request.POST.get('amount')
        item.payment_mode = request.POST.get('payment_mode')
        item.voucher_number = request.POST.get('voucher_number') if request.POST.get('transaction_option') == 'voucher' else None
        item.evoucher_number = request.POST.get('evoucher_number')

        # Handle file uploads
        item.bill_photo = request.FILES.get('bill_photo', item.bill_photo)
        item.gst_photo = request.FILES.get('gst_photo', item.gst_photo)
        item.proof_photo = request.FILES.get('proof_photo', item.proof_photo)

        item.save()

        return redirect('item_form')
    
import pytesseract
from pytesseract import Output
from PIL import Image, ImageEnhance
import cv2
import numpy as np
import re
import matplotlib.pyplot as plt
plt.rcParams["figure.figsize"] = (20, 30)



from django.core.files.storage import FileSystemStorage
from django.http import JsonResponse

def upload_temporary(request):
    if request.method == 'POST' and request.FILES.get('file'):
        uploaded_file = request.FILES['file']
        fs = FileSystemStorage(location='media/tmp/')  # Specify a temporary folder
        filename = fs.save(uploaded_file.name, uploaded_file)
        file_url = fs.url(filename)
        
        return JsonResponse({'success': True, 'file_url': file_url})

    return JsonResponse({'success': False, 'message': 'No file uploaded.'})




from decimal import Decimal, InvalidOperation
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db.models import Sum
import json
import logging

from .models import User, AdvanceGroup, AdvanceAssignment, Expense, AdvanceGroupUpdateLog

logger = logging.getLogger(__name__)

from decimal import Decimal, InvalidOperation
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db.models import Sum
import json
import logging
from .models import User, AdvanceGroup, AdvanceAssignment, Expense, AdvanceGroupUpdateLog
from django.core.files.base import ContentFile
from django.db import transaction as db_transaction

logger = logging.getLogger(__name__)

from django.core.files.base import ContentFile

def assign_advances(request):
    from django.utils import timezone
    print("\n================ ASSIGN_ADVANCES CALLED ================")

    selected_group = None
    show_balance = False
    show_manage = False
    show_history = False

    # <<< FIX: ensure allocation variable always exists >>>
    allocation = None

    if request.method == "POST":
        print("POST DATA:", request.POST)
        print("FILES:", request.FILES)

        if "new_amount" in request.POST:
            print("➡️ Redirecting due to new_amount (balance update)")
            return redirect("assign_advances")

        group_name = request.POST.get("group_name", "").strip()
        user_ids = request.POST.getlist("users")
        amount_raw = request.POST.get("advance_amount", "0")
        proof_file = request.FILES.get("proof_file")
        leader_id = request.POST.get("leader")
        payment_mode = request.POST.get("payment_mode", "net_banking")

        print(f"➡️ Group: {group_name}")
        print(f"➡️ Users: {user_ids}")
        print(f"➡️ Leader: {leader_id}")
        print(f"➡️ Payment Mode: {payment_mode}")
        print(f"➡️ Advance Amount Raw: {amount_raw}")

        try:
            amount = Decimal(amount_raw)
        except InvalidOperation:
            print("❌ Invalid amount")
            messages.error(request, "Invalid amount.")
            return redirect("assign_advances")

        # Create / Get Group
        with db_transaction.atomic():

            # LOCK the group row (or future row)
            group = AdvanceGroup.objects.select_for_update().filter(
                name__iexact=group_name
            ).first()

            if not group:
                group = AdvanceGroup.objects.create(
                    name=group_name,
                    total_advance=0,
                    payment_mode=payment_mode
                )
            else:
                group.payment_mode = payment_mode
                group.save()

            previous_amount = group.total_advance

            # HARD BACKEND BLOCK
            if previous_amount > 0:
                messages.error(
                    request,
                    f"Group '{group.name}' already has ₹{previous_amount}. "
                    "Use Update Balance to add more."
                )
                return redirect("assign_advances")

            # FIRST ASSIGN ONLY
            group.total_advance = amount
            group.save()

        print("➡️ Previous Total:", previous_amount)
        print("➡️ Updated Total:", group.total_advance)

        # Log update
        AdvanceGroupUpdateLog.objects.create(
            group=group,
            previous_amount=previous_amount,
            new_amount=group.total_advance,
            updated_by=request.user,
            proof_file=proof_file,
        )
        print("➡️ Update log saved")

        # Assign users & leader
        success, failed = 0, 0
        for uid in user_ids:
            try:
                user = User.objects.get(id=uid)
                is_leader = str(uid) == leader_id

                if is_leader:
                    print("➡️ Setting Leader:", user.username)
                    AdvanceAssignment.objects.filter(group=group, is_leader=True).update(is_leader=False)
                    group.leader = user
                    group.save()

                AdvanceAssignment.objects.update_or_create(
                    user=user, group=group, defaults={"is_leader": is_leader}
                )
                print("✔️ Assigned user:", user.username)
                success += 1

            except Exception as e:
                print("❌ Failed assigning user:", e)
                failed += 1

        # Clone proof file
        proof_copy = None
        if proof_file:
            print("➡️ Cloning proof file")
            proof_file.seek(0)
            proof_copy = ContentFile(proof_file.read(), name=proof_file.name)

        print("\n====== ALLOCATION EXPENSE UPDATE START ======")

        allocation_qs = Expense.objects.filter(
            advance_group=group,
            is_allocation=True
        )

        print("➡️ Existing Allocation Count:", allocation_qs.count())

        if allocation_qs.exists():
            print("➡️ Updating existing allocation expenses")

            # choose first allocation as 'allocation' reference for context
            allocation = allocation_qs.first()

            for allocation_expense in allocation_qs:
                print(f"   Updating Allocation ID: {allocation_expense.id}")

                allocation_expense.amount = group.total_advance
                allocation_expense.payment_mode = group.payment_mode
                allocation_expense.transaction_date = timezone.now().date()
                allocation_expense.item_name = group.name

                if proof_copy:
                    print("   ➡️ Adding new ProofPhoto for allocation expense")
                    ProofPhoto.objects.create(
                        expense=allocation_expense,
                        file=proof_copy
                    )

                if not allocation_expense.evoucher_number:
                    allocation_expense.evoucher_number = generate_evoucher_number()
                    print("   ➡️ Assigned new EV Number")

                allocation_expense.save()
                print("   ✔️ Saved updated allocation")

        else:
            print("➡️ No allocation exists → Creating new one")

            allocation = Expense.objects.create(
                created_by=request.user,
                item_type="advances",
                transaction_option="voucher",
                transaction_category="internal",
                internal_option=group.leader.get_full_name() if group.leader else "N/A",
                transaction_details=f"Group leader: {group.leader.get_full_name() if group.leader else 'N/A'}",
                amount=group.total_advance,
                transaction_date=group.created_at.date(),
                item_name=group.name,
                payment_mode=group.payment_mode,
                advance_group=group,
                evoucher_number=generate_evoucher_number(),
                is_allocation=True,
                is_approved=True,
                is_rejected=False
            )
            if proof_file:
                proof_file.seek(0)
                fresh_copy = ContentFile(
                    proof_file.read(),
                    name=f"{uuid.uuid4()}_{proof_file.name}"
                )
                ProofPhoto.objects.create(expense=allocation, file=fresh_copy)
            print("✔️ Created new allocation expense")

        print("====== ALLOCATION EXPENSE UPDATE END ======\n")

        messages.success(
            request,
            f"₹{amount} assigned to group '{group.name}'. {success} user(s) assigned. {failed} failed. Expense voucher updated/created.",
        )
        return redirect("assign_advances")

    # ---------------- GET ----------------
    print("➡️ GET request received")

    group_name = request.GET.get("group_name", "").strip()
    action = request.GET.get("action", "")
    print(f"➡️ GET params — group_name: {group_name}, action: {action}")

    assigned_user_ids = []
    leader_id = None
    logs = []

    if group_name:
        selected_group = AdvanceGroup.objects.filter(name=group_name).first()
        print("➡️ Selected Group:", selected_group)

        if selected_group:
            if action == "balance":
                show_balance = True
                print("➡️ Showing balance")
            elif action == "manage":
                show_manage = True
                print("➡️ Managing users")
            elif action == "history":
                show_history = True
                print("➡️ Showing history")

            assigned_user_ids = list(
                AdvanceAssignment.objects.filter(
                    group=selected_group
                ).values_list("user_id", flat=True)
            )
            print("➡️ Assigned Users:", assigned_user_ids)

            leader = AdvanceAssignment.objects.filter(
                group=selected_group, is_leader=True
            ).first()
            leader_id = leader.user.id if leader else None
            print("➡️ Leader ID:", leader_id)

            used_total = (
                Expense.objects.filter(
                    item_type="advances",
                    advance_group=selected_group,
                    is_allocation=False,
                    is_draft=False,
                    status__in=["pending", "paid"],
                ).aggregate(total=Sum("amount"))["total"]
                or Decimal("0.00")
            )

            remaining_total = selected_group.total_advance - used_total
            if remaining_total < 0:
                remaining_total = Decimal("0.00")

            print("➡️ Used:", used_total)
            print("➡️ Remaining:", remaining_total)

            selected_group.used_advance = used_total
            selected_group.remaining_balance = remaining_total
            selected_group.save()

            logs = list(
                AdvanceGroupUpdateLog.objects.filter(group=selected_group).order_by("timestamp")
            )
            print("➡️ Logs found:", len(logs))

            # Also set allocation for GET if exists (so template won't crash)
            allocation = Expense.objects.filter(advance_group=selected_group, is_allocation=True).first()

    first_log = logs[-1] if logs else None

    print("➡️ Rendering page now...\n")

    context = {
        "users": User.objects.all(),
        "groups": AdvanceGroup.objects.filter(is_closed=False),
        "selected_group": selected_group,
        "assigned_user_ids": assigned_user_ids,
        "leader_id": leader_id,
        "show_balance": show_balance,
        "show_manage": show_manage,
        "show_history": show_history,
        "logs": sorted(logs, key=lambda log: log.timestamp, reverse=True),
        "first_log": first_log,
        "assigned_user_ids_map": {
            group.id: list(
                AdvanceAssignment.objects.filter(group=group).values_list("user_id", flat=True)
            )
            for group in AdvanceGroup.objects.all()
        },
        "groups_json": json.dumps({str(group.id): group.name for group in AdvanceGroup.objects.all()}),

        # <<< safe: ensure the template key exists even if no allocation >>> 
        "expenses": allocation,
    }

    return render(request, "xp/assign_advances.html", context)


from datetime import datetime

def get_date_range(request):
    """
    Reads ?start_date=YYYY-MM-DD&end_date=YYYY-MM-DD and returns (start_date, end_date) as date objects.
    Returns (None, None) if not provided or invalid.
    """
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    try:
        if start_date:
            start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
        if end_date:
            end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
    except ValueError:
        return None, None

    return start_date, end_date


from decimal import Decimal
from django.db.models import Sum
from django.shortcuts import render
from .models import AdvanceGroup, AdvanceAssignment, Expense
from django.contrib.auth.models import User


from decimal import Decimal
from django.db.models import Sum
from django.shortcuts import render
from .models import AdvanceGroup, AdvanceAssignment, Expense
from django.contrib.auth.models import User

def calculate_group_balances(group):
    """
    Return (used, remaining, overdue)

    used = total expenses in this group
    remaining = total_advance - used  (can be negative!)
    overdue = absolute negative amount
    """
    used = (
        Expense.objects.filter(
            item_type="advances",
            advance_group=group,
            is_allocation=False,
            is_draft=False,
            status__in=["pending", "approved", "paid"]
        ).aggregate(total=Sum("amount"))["total"]
        or Decimal("0.00")
    )

    remaining = group.total_advance - used   # allow negative
    overdue = abs(remaining) if remaining < 0 else Decimal("0.00")

    return used, remaining, overdue

def view_assigned_groups(request):
    selected_group_name = request.GET.get("group_name")
    selected_user_id = request.GET.get("user_id")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    group_data = []

    groups_qs = AdvanceGroup.objects.all().order_by("-created_at")
    dropdown_groups = AdvanceGroup.objects.filter(is_closed=False).order_by("-created_at")

    if not request.user.is_staff and not request.user.is_superuser:
        groups_qs = groups_qs.filter(
            id__in=AdvanceAssignment.objects.filter(user=request.user).values_list("group_id", flat=True)
        )

    if start_date and end_date:
        # Accept strings; if invalid they will raise and be ignored
        try:
            s = datetime.strptime(start_date, "%Y-%m-%d").date()
            e = datetime.strptime(end_date, "%Y-%m-%d").date()
            groups_qs = groups_qs.filter(created_at__date__gte=s, created_at__date__lte=e)
        except Exception:
            pass

    if selected_group_name:
        groups_qs = groups_qs.filter(name=selected_group_name)

    if selected_user_id:
        groups_qs = groups_qs.filter(
            id__in=AdvanceAssignment.objects.filter(user_id=selected_user_id).values_list("group_id", flat=True)
        )

    for group in groups_qs:
        assigned_user_ids = list(
            AdvanceAssignment.objects.filter(group=group).values_list("user_id", flat=True)
        )
        assigned_users = User.objects.filter(id__in=assigned_user_ids)

        if not request.user.is_staff and request.user.id not in assigned_user_ids:
            continue

        user_rows = {}

        for user in assigned_users:
            user_rows[user.id] = {
                "user": user,
                "credit": group.total_advance,
                "rows": [],
                "balance": group.total_advance,
            }

        expenses = Expense.objects.filter(
            item_type="advances",
            advance_group=group,
            status__in=["pending", "paid"],
            is_draft=False,
            is_allocation=False,
        ).order_by("date", "id")

        if start_date and end_date:
            try:
                s = datetime.strptime(start_date, "%Y-%m-%d").date()
                e = datetime.strptime(end_date, "%Y-%m-%d").date()
                expenses = expenses.filter(date__range=[s, e])
            except Exception:
                pass

        for exp in expenses:
            user = exp.created_by
            if not user:
                continue
            if user.id not in user_rows:
                user_rows[user.id] = {
                    "user": user,
                    "credit": group.total_advance,
                    "rows": [],
                    "balance": group.total_advance,
                }

            debit = exp.amount or Decimal("0.00")
            balance = user_rows[user.id]["balance"] - debit
            user_rows[user.id]["balance"] = balance

            user_rows[user.id]["rows"].append({
                "item": exp.item_name or "Unnamed",
                "credit": user_rows[user.id]["credit"],
                "debit": debit,
                "remaining": balance,
                "date": exp.transaction_date or exp.date,
                "overdue": balance < 0,
                "overdue_amount": abs(balance) if balance < 0 else None,
            })

        used, remaining, overdue = calculate_group_balances(group)

        # cleared_amount = Decimal("0.00")
        # if group.remaining_balance == 0 and remaining < 0:
        #     cleared_amount = abs(remaining)

        cleared_amount = Decimal("0.00")

        # show cleared ONLY if a PAID expense exists for this group
        has_paid_expense = Expense.objects.filter(
            advance_group=group,
            status="paid",
            is_allocation=False,
            is_draft=False,
        ).exists()

        if has_paid_expense and remaining >= 0 and overdue > 0:
            cleared_amount = overdue

        initial_proof = group.proof_file.url if hasattr(group, "proof_file") and group.proof_file else None
        update_proofs = [log.proof_file.url for log in group.update_logs.all() if log.proof_file]

        group_data.append(
            {
                "group": group,
                "used_advance": used,
                "remaining_balance": group.remaining_balance,   # <-- DB value
                "calculated_remaining": remaining,              # <-- may be negative
                "overdue_amount": overdue,
                "cleared_amount": cleared_amount,
                "users": list(user_rows.values()),
                "initial_proof": initial_proof,
                "update_proofs": update_proofs,
                "assigned_user_ids": assigned_user_ids,
            }
        )

    all_users = User.objects.all()

    return render(
        request,
        "xp/view_assigned_groups.html",
        {
            "group_data": group_data,
            "groups": groups_qs,
            "selected_group": selected_group_name,
            "selected_user": selected_user_id,
            "all_users": all_users,
            "day_ranges": [7, 15, 30],
            "dropdown_groups": dropdown_groups,
        },
    )


# utils.py

from datetime import datetime
from decimal import Decimal
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .models import Expense, AdvanceGroup


def get_filter_params(request):
    """Extracts start_date, end_date, group_id from request.GET."""
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    group_id = request.GET.get("group_id")

    if start_date:
        start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
    if end_date:
        end_date = datetime.strptime(end_date, "%Y-%m-%d").date()

    return start_date, end_date, group_id


@login_required
def export_advances_excel(request):
    start_date, end_date, group_id = get_filter_params(request)
    group_id_int = int(group_id) if group_id and group_id.isdigit() else None

    # Groups queryset
    groups_qs = AdvanceGroup.objects.all().order_by("name")
    if group_id_int:
        groups_qs = groups_qs.filter(id=group_id_int)

    # Apply date filter on group creation date
    if start_date and end_date:
        groups_qs = groups_qs.filter(created_at__date__range=[start_date, end_date])

    # Excel workbook setup
    wb = Workbook()
    ws = wb.active
    ws.title = "Advances"

    # Base header
    base_header = ["USER", "ITEM", "CREDIT", "DEBIT", "REMAINING", "DATE"]
    ws.append(base_header)

    row_idx = 2

    for group in groups_qs:
        group_expenses = Expense.objects.filter(
            advance_group=group,
            item_type="advances",
            status__in=["pending", "paid"],
            is_draft=False,
        ).order_by("date", "id")

        if not group_expenses.exists():
            continue

        # All proof files for this group
        proof_files = [
            request.build_absolute_uri(log.proof_file.url)
            for log in group.update_logs.all()
            if log.proof_file
        ]

        balance_map = {}

        for exp in group_expenses:
            user = exp.created_by
            if not user:
                continue

            # Track balance per user
            if user.id not in balance_map:
                balance_map[user.id] = group.total_advance or Decimal("0.00")

            debit = exp.amount or Decimal("0.00")
            remaining = balance_map[user.id] - debit
            balance_map[user.id] = remaining

            # Base row data
            row_data = [
                f"{user.get_full_name() or user.username} ({group.name})",
                exp.item_name or "Unnamed",
                f"₹{(group.total_advance or 0):.2f}",
                f"₹{debit:.2f}",
                f"₹{remaining:.2f}",
                exp.date.strftime("%B %d, %Y") if exp.date else "-",
            ]

            # Add proof columns
            for proof_url in proof_files:
                row_data.append(proof_url)

            # Write row
            ws.append(row_data)

            # Add clickable hyperlinks for proofs
            col_offset = len(base_header)  # starting column index for proofs
            for i, proof_url in enumerate(proof_files):
                cell = ws.cell(row=row_idx, column=col_offset + i + 1)
                cell.hyperlink = proof_url
                cell.value = f"Proof {i+1}"
                cell.style = "Hyperlink"

            row_idx += 1

    # Auto-adjust column width
    for i, col in enumerate(ws.columns, 1):
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(i)].width = max_len + 3

    # Return Excel file
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="advances.xlsx"'
    wb.save(response)
    return response

from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle

def export_advances_pdf(request):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = [Paragraph("Advance Report Summary", styles['Title']), Spacer(1, 12)]

    data = [['Group', 'User', 'Total', 'Used', 'Remaining']]
    for group in AdvanceGroup.objects.all().order_by('name'):
        for assignment in AdvanceAssignment.objects.filter(group=group):
            used = Expense.objects.filter(advance_group=group, created_by=assignment.user).aggregate(total=Sum('amount'))['total'] or 0
            data.append([
                group.name,
                assignment.user.get_full_name() or assignment.user.username,
                f"₹{group.total_advance:.2f}",
                f"₹{used:.2f}",
                f"₹{group.total_advance - used:.2f}",
            ])
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename="advance_report.pdf")

@require_POST
def update_group_balance(request, group_id):
    print("\n================ UPDATE_GROUP_BALANCE CALLED ================")
    group = get_object_or_404(AdvanceGroup, id=group_id)

    print(f"➡️ Updating group: {group.name} (ID: {group_id})")

    try:
        additional_amount = Decimal(request.POST.get("new_amount"))
        print(f"➡️ Additional Amount: {additional_amount}")
    except (InvalidOperation, TypeError):
        print("❌ Invalid amount")
        messages.error(request, "Invalid amount entered.")
        return redirect("assign_advances")

    if additional_amount <= 0:
        print("❌ Amount must be > 0")
        messages.error(request, "Amount must be greater than zero.")
        return redirect("assign_advances")

    prev = group.total_advance
    group.total_advance += additional_amount
    group.save()
    print(f"➡️ Previous Total: {prev}, Updated Total: {group.total_advance}")

    # Recalculate used, remaining, overdue
    used, remaining, overdue = calculate_group_balances(group)

    # Save new balances
    group.used_advance = used
    group.remaining_balance = remaining
    group.save()

    print(f"➡️ Used: {used}, Remaining: {remaining}, Overdue: {overdue}")

    proof_file = request.FILES.get("proof_file")

    # Log the update
    AdvanceGroupUpdateLog.objects.create(
        group=group,
        previous_amount=prev,
        new_amount=group.total_advance,
        updated_by=request.user,
        proof_file=proof_file
    )
    print("➡️ Update log saved")

    print("====== ALLOCATION EXPENSE UPDATE CHECK ======")

    # Get allocation voucher
    allocation_expense = Expense.objects.filter(
        advance_group=group,
        is_allocation=True
    ).first()

    if allocation_expense:
        print(f"➡️ Found allocation expense: ID {allocation_expense.id}")

        allocation_expense.amount = group.total_advance
        allocation_expense.transaction_date = timezone.now().date()
        allocation_expense.item_name = group.name

        if hasattr(group, "payment_mode"):
            print(f"➡️ Updating payment mode to: {group.payment_mode}")
            allocation_expense.payment_mode = group.payment_mode

        # If new proof uploaded → add a new ProofPhoto (append, don't overwrite)
        if proof_file:
            proof_file.seek(0)
            proof_copy = ContentFile(
                proof_file.read(),
                name=f"{uuid.uuid4()}_{proof_file.name}"
            )
            ProofPhoto.objects.create(expense=allocation_expense, file=proof_copy)

        allocation_expense.save()
        print("✔️ Allocation expense UPDATED successfully")

    else:
        print("❌ No existing allocation expense found! Nothing to update.")

    messages.success(
        request,
        f"Added ₹{additional_amount} to group '{group.name}'. "
        f"New total: ₹{group.total_advance}, Remaining: ₹{group.remaining_balance}"
    )

    print("================ UPDATE COMPLETE ================\n")
    return redirect("assign_advances")



from .models import Expense, CashVoucher, User, BorrowedAmount
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from datetime import date, datetime
import decimal
from decimal import Decimal, InvalidOperation
from django.shortcuts import redirect
from django.contrib import messages
from website.models import CashVoucher
from .models import AdvanceAssignment

def new_item_form(request, user_id):
    print("hgo")
    current_date = date.today()
    selected_user = get_object_or_404(User, id=user_id)
    user_groups = AdvanceAssignment.objects.filter(
    user=request.user,
    group__is_closed=False
).select_related('group') 
    is_draft = False  # Initialize is_draft for all request methods
    total_amount = 0  # Initialize total_amount for all request methods
    transaction_category = ""
    conveyance_options = Conveyance.objects.all()
    
    if request.method == "POST":
        print(f"post",request.POST)


        item_type = request.POST.get("item_type")
        item_names = request.POST.getlist('item_name', [])
        item_name = next((name for name in item_names if name.strip()), None)
        transaction_option = request.POST.get("transaction_option", "").strip().lower()
        print(transaction_option)
        transaction_category = request.POST.get("transaction_category")
        kilometers = request.POST.get("kilometers")  # New field
        print(f"Kilometers received: {kilometers}")
        vehicle_type = request.POST.get("vehicle_type") 
        internal_options = request.POST.getlist("internal_option")
        external_type = request.POST.get("external_type")
        total_amount = request.POST.get("amount", "").replace(",", "").strip()
        print(f"ttamt",total_amount)


        
        payment_mode = request.POST.get("payment_mode")
        voucher_number = request.POST.get("voucher_number")
        evoucher_number = request.POST.get("e_voucher_number")
        transaction_dates = request.POST.getlist('transaction_date', [])
        transaction_date = next(
                (datetime.strptime(date.strip(), "%Y-%m-%d").date() for date in transaction_dates if date.strip()),
                None,
            )
        bill_photo = request.FILES.get("bill_photo")
        gst_photo = request.FILES.get("gst_photo")
     
        proof_files = request.FILES.getlist("proof_photos")
        if transaction_option in ["2_wheeler", "4_wheeler"] and kilometers:
            print(f"Kilometers provided: {kilometers}")

            try:
                # Try to convert kilometers to float and log the conversion
                kilometers = Decimal(kilometers)
                print(f"Kilometers converted to float: {kilometers}")
            
                
                # Fetch conveyance information based on vehicle_type
                print(f"toption",transaction_option)
                vehicle_type = transaction_option
                print(f"vhtype",vehicle_type)
                conveyance = Conveyance.objects.filter(vehicle_type=vehicle_type).first()
                print(f"Conveyance fetched: {conveyance}")
                
                if conveyance:
                    # Calculate total_amount if conveyance exists
                    print(f"if convey",total_amount)
                    total_amount = kilometers * conveyance.price_per_km
                    print(f"Calculated amount for {vehicle_type}: {total_amount}")
                else:
                    # If no conveyance found, show an error
                    messages.error(request, f"No conveyance rate found for {vehicle_type}.")
                    total_amount = 0

            except ValueError:
                # If conversion fails, show an error and reset total_amount to 0
                messages.error(request, "Invalid kilometers value.")
                total_amount = 0
                print("Error: Invalid kilometers value. Setting total_amount to 0.")
        
            # File uploads
            
        # Check if 'draft_status' is in the POST data and set is_draft accordingly

        draft_statuss = request.POST.get('draft_status') 
        print(f"defmain",draft_statuss) # Default to 'false' if not provided
        draft_status = request.POST.get('draft_status', 'false')  # Default to 'false' if not provided
        print(f"def",draft_status)
        is_draft = draft_status.lower() == 'true'  # Set is_draft to True if 'draft_status' is 'true', otherwise False
        print(f"before storing to model",total_amount)
        # Deduct from AdvanceGroup if item_type is 'advances'
        total_amount_decimal = Decimal(str(total_amount or 0))

        selected_group_id = request.POST.get("selected_group_id")
        if item_type == 'advances':
            try:
                group = AdvanceGroup.objects.get(id=selected_group_id)
                if not AdvanceAssignment.objects.filter(user=request.user, group=group).exists():
                    messages.error(request, "You are not assigned to this group.")
                    return redirect("new_item_form", user_id=user_id)

                # Deduct from group after validation
                group.used_advance += total_amount_decimal

                # ✅ Recalculate remaining balance
                used = Expense.objects.filter(
                    item_type="advances",
                    advance_group=group,
                    status__in=["pending", "paid"],
                    is_draft=False,
                    is_allocation=False ,
                ).aggregate(total=Sum("amount"))["total"] or Decimal("0.00")

                group.remaining_balance = max(group.total_advance - used, Decimal("0.00"))
                group.save(update_fields=["used_advance", "remaining_balance"])

            except AdvanceGroup.DoesNotExist:
                messages.error(request, "Selected group not found.")
                return redirect("new_item_form", user_id=user_id)

        
        # Auto-set status to 'paid' if item_type is 'advances'
        # Save the primary expense for User 1
        # Determine if this expense requires approval
        needs_approval = False

        # Any non-conveyance, non-voucher item with no proof will need approval
        # Determine if approval is needed
        item_type = request.POST.get("item_type", "").strip().lower()
        transaction_option = request.POST.get("transaction_option", "").strip().lower()

        print(f"[DEBUG] item_type = {item_type}")
        print(f"[DEBUG] transaction_option = {transaction_option}")

        bill_or_gst_present = bool(request.FILES.get("bill_photo")) or bool(request.FILES.get("gst_photo"))
        proof_present = bool(request.FILES.get("proof_photo")) or bool(proof_files)

        print(f"[DEBUG] Bill or GST present: {bill_or_gst_present}")
        print(f"[DEBUG] Proof present: {proof_present}")

        # Main logic
        needs_approval = False

        if transaction_option == "voucher":
            print("✅ Voucher condition triggered")
            needs_approval = True
        elif item_type == "conveyance":
            print("✅ Conveyance: No approval needed")
            needs_approval = False
        elif bill_or_gst_present and proof_present:
            print("✅ Both Bill/GST and Proof present: No approval")
            needs_approval = False
        else:
            print("❌ Approval needed")
            needs_approval = True

        print(f"[DEBUG] Final needs_approval = {needs_approval}")

        # Determine status
        if item_type == "advances" and not needs_approval:
            status = "paid"  # Auto-pay if no approval needed
        else:
            status = "pending"  # Default to pending       
        expense = Expense.objects.create(
            
            created_by=request.user,
            item_type=item_type,
            item_name=item_name,
            transaction_option=transaction_option,

            transaction_category=transaction_category,
            internal_option=",".join(internal_options) if transaction_category == "internal" else None,
            external_type=external_type if transaction_category == "external" else None,
            
            payment_mode=payment_mode,
            voucher_number=voucher_number if transaction_option == "voucher" else None,
            evoucher_number=evoucher_number,
            bill_photo=bill_photo,
            gst_photo=gst_photo,
            km=kilometers if transaction_option in ["2_wheeler", "4_wheeler"] else None,
            transaction_date=transaction_date,
            is_draft=is_draft,  # Set the is_draft field based on the form submission
            amount=total_amount,
            advance_group=group if item_type == 'advances' else None,
            status=status,
        )
        
        for file in proof_files:
            if file:
                ProofPhoto.objects.create(expense=expense, file=file)

        print(f"after context",transaction_option)

        print(f"isdraft",is_draft)
        # Handle borrowed amounts and replicate vouchers for multiple borrowed users
                # Handle borrowed amounts and replicate vouchers for multiple borrowed users
        borrowed_amounts = request.POST.getlist("borrowed_amounts[]")
        borrowed_froms = request.POST.getlist("borrowed_froms[]")

        # Initialize remaining amount
        remaining_amount = decimal.Decimal(total_amount)

        for borrowed_amount, borrowed_from in zip(borrowed_amounts, borrowed_froms):
            if borrowed_amount and borrowed_from:
                borrowed_amount = decimal.Decimal(borrowed_amount)  # Convert borrowed amount to Decimal

                if borrowed_amount <= remaining_amount:
                    # Create a BorrowedAmount entry
                    BorrowedAmount.objects.create(
                        expense=expense,
                        amount=borrowed_amount,
                        borrowed_from_id=borrowed_from,
                    )

                    # Replicate the voucher for the borrowed user (User 2, User 3, etc.)
                    Expense.objects.create(
                        created_by=User.objects.get(id=borrowed_from),
                        item_type=item_type,
                        item_name=f"Borrowed - {item_name}",
                        transaction_option=transaction_option,
                        internal_option=request.user,
                        transaction_category="internal",
                        amount=borrowed_amount,  # Assign borrowed amount
                        payment_mode=payment_mode,
                        voucher_number=voucher_number if transaction_option == "voucher" else None,
                        evoucher_number=evoucher_number,
                        bill_photo=None,
                        gst_photo=None,
                        proof_photo=None,
                        transaction_date=transaction_date,
                    )

                    # Reduce the remaining amount
                    remaining_amount -= borrowed_amount
                else:
                    continue  # Skip if borrowed amount exceeds available funds

        # Ensure User 1’s expense is updated with the correct remaining amount
        expense.amount = remaining_amount  # Should be 0 if fully borrowed
        expense.save()

        # Initialize cvamount before the block
        cvamount = 0

        if transaction_category == "external":
            cvamount = request.POST.get('cvamount', '').strip()
            if cvamount:
                    try:
                        cvamount = decimal.Decimal(cvamount)
                    except decimal.InvalidOperation:
                        cvamount = 0  # Handle invalid decimals gracefully
            else:
                    cvamount = 0  

        if "tips_checkbox" in request.POST:
            is_tips_selected = request.POST.get("tips_checkbox") == "on"

            if is_tips_selected:
        # Fetch the Cash Voucher details from the form
                cv_amount = request.POST.get("cvamount", "").strip()
                paid_to = request.POST.get("paid_to", "").strip()
                cv_item_name = request.POST.get("item_name", "").strip()
                cv_transaction_date = request.POST.get("transaction_date", "").strip()
                cv_voucher_number = request.POST.get("tip_voucher_number", "")
                tip_proof_photo = request.FILES.get("proof_photo") 

                # Validate all required fields
                if not (cv_amount and paid_to and cv_item_name and cv_transaction_date):
                    messages.error(request, "All Cash Voucher fields are required.")
                    return redirect("item_form")

                # Validate and convert cv_amount to Decimal
                try:
                    cv_amount = Decimal(cv_amount)
                except InvalidOperation:
                    messages.error(request, "Invalid Cash Voucher amount.")
                    return redirect("item_form")

                # Create the Cash Voucher with 'pending' status
                cash_voucher = CashVoucher.objects.create(
                    created_by=request.user,
                    amount=cv_amount,
                    paid_to=paid_to,
                    item_name=cv_item_name,
                    transaction_date=cv_transaction_date,
                    voucher_number=cv_voucher_number,
                    status="pending",  # Set status to 'pending'
                    expense=expense,
                    proof_photo=tip_proof_photo,
                )
                messages.success(request, "Cash Voucher saved successfully! Awaiting approval.")

                # Link Cash Voucher amount to the E-Voucher only if voucher is approved
                evoucher_number = request.POST.get("e_voucher_number", "").strip()
                if evoucher_number:
                    try:
                        # Find the corresponding Expense with the given eVoucher number
                        evoucher_expense = Expense.objects.get(evoucher_number=evoucher_number)

                        # Only update the amount if the cash voucher is approved
                        if cash_voucher.status == 'approved':
                            evoucher_expense.amount += cv_amount
                            evoucher_expense.save()
                            messages.success(request, "Cash Voucher amount added to E-Voucher successfully!")
                        else:
                            messages.error(request, "Cash Voucher is not approved yet. Cannot update the E-Voucher amount.")
                    except Expense.DoesNotExist:
                        messages.error(request, f"No Expense found with E-Voucher number {evoucher_number}.")
                else:
                    messages.error(request, "E-Voucher number is required to link the Cash Voucher.")
            else:
                messages.error(request, "Please complete all required Cash Voucher fields.")

        # Redirect back to the draft page or same form to review the draft data
        if is_draft:
            return redirect("draft_vouchers")  # Redirect to the draft vouchers list page
        return redirect("item_form")  # Ensure 'item_form' is a valid URL pattern name

    # Fetch approved vouchers for the selected user
    used_voucher_numbers = list(Expense.objects.values_list("voucher_number", flat=True).exclude(voucher_number__isnull=True).distinct())
    approved_vouchers = CashVoucher.objects.filter(
        created_by=selected_user, status="approved"
    ).exclude(voucher_number__in=used_voucher_numbers)
    users = User.objects.all()

    evoucher_number = generate_evoucher_number()
    submitted_data = request.POST or None
    cv_voucher_number =generate_voucher_number()
    remaining_balance = None
    if request.user.is_authenticated:
        try:
            assignments = AdvanceAssignment.objects.filter(user=request.user).select_related("group")
            if assignments.exists():
                # Just pick the first group's balance to display by default (or enhance logic as needed)
                group = assignments.first().group
                remaining_balance = group.remaining_balance
            else:
                remaining_balance = None
        except AdvanceAssignment.DoesNotExist:
            remaining_balance = None

    context = {
        "approved_vouchers": approved_vouchers,
        "evoucher_number": evoucher_number,
        "current_date": current_date,
        "selected_user": selected_user,
        "users": users,
        'conveyance_options': conveyance_options,
        'submitted_data': submitted_data,
        'cv_voucher_number':cv_voucher_number,
        'remaining_balance': remaining_balance,
        'user_groups': user_groups,
        
    }

    return render(request, "xp/newitemform.html", context)


from decimal import Decimal
from django.views.decorators.http import require_POST
from django.http import JsonResponse
from .models import CashVoucher, Expense

from django.db.models import F

@require_POST
def approve_tip_voucher(request):
    print("got in")
    try:
        voucher_id = request.POST.get('voucher_id')
        expense_id = request.POST.get('expense_id')

        voucher = CashVoucher.objects.get(id=voucher_id)
        expense = Expense.objects.get(id=expense_id)

        if voucher.status != 'pending':
            return JsonResponse({'success': False, 'error': 'Already approved or rejected'})

        voucher.status = 'approved'
        voucher.save()

        # Use F expression to add directly in DB
        if voucher.amount:
            Expense.objects.filter(id=expense.id).update(amount=F('amount') + voucher.amount)

        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


from django.views.decorators.http import require_POST
from django.shortcuts import redirect, get_object_or_404
from .models import AdvanceGroup
from django.contrib import messages

@require_POST
def delete_group(request, group_id):
    group = get_object_or_404(AdvanceGroup, id=group_id)
    group_name = group.name
    group.delete()
    messages.success(request, f"Group '{group_name}' has been deleted.")
    return redirect('view_assigned_groups')  # Update this to your group listing URL name




from django.contrib.auth.decorators import login_required
from .models import Expense
from django.shortcuts import render

@login_required
def draft_vouchers(request):
    user = request.user

    if user.is_staff or user.is_superuser:
        # ✅ Admin: see all drafts
        draft_vouchers = Expense.objects.filter(is_draft=True)
    else:
        # ✅ User: see only their drafts
        draft_vouchers = Expense.objects.filter(
            is_draft=True,
            created_by=user
        )

    context = {
        "expenses": draft_vouchers
    }

    return render(request, "xp/draft_vouchers.html", context)



from django.contrib.auth.models import User
from django.views.decorators.http import require_POST
from django.shortcuts import get_object_or_404, redirect
from .models import AdvanceGroup, AdvanceAssignment
from django.contrib import messages

from django.views.decorators.http import require_POST
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from .models import AdvanceGroup, AdvanceAssignment

@require_POST
def update_group_members(request, group_id):
    group = get_object_or_404(AdvanceGroup, id=group_id)
    selected_user_ids = request.POST.getlist("users")
    selected_user_ids = set(map(int, selected_user_ids))

    # a unselected members
    AdvanceAssignment.objects.filter(group=group).exclude(user_id__in=selected_user_ids).delete()

    # Add or update members
    for uid in selected_user_ids:
        AdvanceAssignment.objects.get_or_create(user_id=uid, group=group)

    # ✅ Update leader in AdvanceAssignment
    leader_id = request.POST.get("leader")
    if leader_id:
        leader_id = int(leader_id)
        if leader_id in selected_user_ids:
            # clear old leader
            AdvanceAssignment.objects.filter(group=group, is_leader=True).update(is_leader=False)
            # set new leader
            AdvanceAssignment.objects.filter(group=group, user_id=leader_id).update(is_leader=True)
        else:
            messages.warning(request, "Leader must be one of the selected members.")
    # ⚠️ if no leader posted → old leader stays

    messages.success(request, f"Updated members for group '{group.name}'.")
    return redirect("assign_advances")




from datetime import datetime
from django.shortcuts import render, get_object_or_404
from .models import CashVoucher, User
from django.db.models import Q
from django.utils.dateparse import parse_date

def cash_voucher(request):
    approved_vouchers = CashVoucher.objects.filter(status='approved').order_by('-id')
    rejected_vouchers = CashVoucher.objects.filter(status='rejected').order_by('-id')
    total_amount = sum(voucher.amount for voucher in approved_vouchers)

    # Get the current date
    current_date = datetime.today().date()

    # Default queryset for vouchers
    cash_vouchers = CashVoucher.objects.all().order_by('-id')

    # Check if the user is admin or superuser
    if not (request.user.is_staff or request.user.is_superuser):
        # Regular users can only see their own vouchers
        cash_vouchers = CashVoucher.objects.filter(created_by=request.user).order_by('-id')

    # User filter logic (from payable view)
    selected_user = None
    user_id = request.GET.get('user_id')  # Get the selected user from the GET request
    if user_id:
        selected_user = User.objects.get(id=user_id)
        cash_vouchers = cash_vouchers.filter(created_by=selected_user)

    # Get the search query and date filters from GET parameters
    search_query = request.GET.get('q', '')
    date_from = request.GET.get('transaction_date_from', '')
    date_to = request.GET.get('transaction_date_to', '')

    # Apply search query filter if it exists
    if search_query:
        cash_vouchers = cash_vouchers.filter(
            Q(voucher_number__icontains=search_query) |
            Q(paid_to__icontains=search_query) |
            Q(item_name__icontains=search_query) |
            Q(amount__icontains=search_query) |
            Q(status__icontains=search_query)
        )

    # Apply date filters if they exist
    if date_from:
        parsed_date_from = parse_date(date_from)
        if parsed_date_from:
            cash_vouchers = cash_vouchers.filter(transaction_date__gte=parsed_date_from)

    if date_to:
        parsed_date_to = parse_date(date_to)
        if parsed_date_to:
            cash_vouchers = cash_vouchers.filter(transaction_date__lte=parsed_date_to)

    # Handle voucher approval
    verify_id = request.GET.get('verify_id')
    if verify_id:
        try:
            voucher = get_object_or_404(CashVoucher, id=verify_id)
            if voucher.status == 'pending':
                voucher.status = 'approved'
                voucher.approved_by = request.user.username 
                voucher.verified_at = timezone.now()  # Track who approved
                voucher.save()

                # Create a notification for the user who created the voucher
            Notification.objects.create(
                    user=voucher.created_by,
                    title=f"Your Cash Voucher has been approved",
                    message=f"Your voucher with number {voucher.voucher_number} of amount  {voucher.amount} has been approved."
                )
        except Exception as e:
            print(f"Error while verifying voucher: {e}")

    # Handle voucher rejection
    reject_id = request.GET.get('reject_id')
    if reject_id:
        try:
            voucher = get_object_or_404(CashVoucher, id=reject_id)
            if voucher.status == 'pending':  # Only allow rejection of pending vouchers
                voucher.status = 'rejected'
                voucher.approved_by = request.user.username
                voucher.save()

                # Create a notification for the user who created the voucher
            Notification.objects.create(
                    user=voucher.created_by,
                    title=f"{CashVoucher.voucher_number} of amount {CashVoucher.amount} is rejected",
                    message=f"Your voucher with number {CashVoucher.voucher_number} has been rejected."
                )
        except Exception as e:
            print(f"Error while rejecting voucher: {e}")


    # Pass the current date, vouchers, and search query to the template
    return render(request, 'xp/cash_voucher.html', {
        'current_date': current_date,
        'vouchers': cash_vouchers,
        'search_query': search_query,
        'date_from': date_from,
        'date_to': date_to,
        'approved_vouchers': approved_vouchers,
        'rejected_vouchers': rejected_vouchers,
        'total_amount': total_amount,
        'selected_user': selected_user,
        'users': User.objects.all(),  # List of all users for the filter

    })

from django.db.models import Sum

def approved_vouchers(request):
    user = request.user
    
    # If the user is admin, fetch all approved vouchers
    if user.is_staff or user.is_superuser:
        vouchers = CashVoucher.objects.filter(status='approved').order_by('-id')
    else:
        # If not an admin, fetch only the approved vouchers created by the user
        vouchers = CashVoucher.objects.filter(created_by=user, status='approved').order_by('-id')

    # Debug print to check the result of the query


    # Calculate the total amount of the approved vouchers
    total_amount = vouchers.aggregate(total_amount=Sum('amount'))['total_amount'] or 0

    # Prepare context with vouchers and total amount
    context = {'vouchers': vouchers, 'total_amount': total_amount}

    return render(request, 'xp/approved_vouchers.html', context)

def rejected_vouchers(request):
    user = request.user
    if user.is_staff or user.is_superuser:
        rejected_vouchers = CashVoucher.objects.filter(status='rejected')
    else:
         rejected_vouchers = CashVoucher.objects.filter(created_by=user, status='rejected')

    context = {'rejected_vouchers': rejected_vouchers}
    return render(request, 'xp/rejected_vouchers.html', context)


def cash_voucher_success(request):
    return render(request, 'xp/success.html')

from django.shortcuts import render, redirect
from .forms import CashVoucherForm
from .models import CashVoucher

def create_cash_voucher(request):
    current_date = datetime.today().date()
    if request.method == 'POST':
        form = CashVoucherForm(request.POST)

        if form.is_valid():
            transaction_date = form.cleaned_data.get('transaction_date')
            form.save()  # Saves the cash voucher in the database
            return redirect('success')  # Redirect to a success page or another view
    else:
        form = CashVoucherForm()
    
    return render(request, 'xp/create_cash_voucher.html', {'form': form})

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from .models import CashVoucher

def reject_voucher(request):
    if request.method == 'POST':
        voucher_id = request.POST.get('voucher_id')
        rejection_reason = request.POST.get('reason')

        # Ensure voucher exists
        voucher = get_object_or_404(CashVoucher, id=voucher_id)

        # Only reject vouchers that are pending
        if voucher.status == 'pending':
            voucher.status = 'rejected'
            voucher.rejection_reason = rejection_reason  # Assuming you have a 'rejection_reason' field

            voucher.save()

            # Show a success message
            messages.success(request, 'Voucher rejected successfully!')
        else:
            messages.error(request, 'Voucher cannot be rejected at this stage.')

        return redirect('cash_voucher')  # Redirect to the page with the vouchers

from django.shortcuts import render, get_object_or_404, redirect
from .models import CashVoucher


def approve_cash_voucher(request, voucher_number):
    # Check if user has permission to approve
    if not request.user.has_perm('app.can_approve'):
        return redirect('permission_denied')  # Redirect to permission denied page if no permission
    
    # Get the voucher by voucher number
    voucher = get_object_or_404(CashVoucher, voucher_number=voucher_number)

    # If voucher status is already approved, prevent re-approval
    if voucher.status == 'approved':
        return redirect('voucher_already_approved')  # Redirect to a page showing it's already approved

    # Approve the voucher
    voucher.status = 'approved'
    voucher.approved_by = request.user.username  # Set the approver as the current user
    voucher.save()

    return redirect('approval_success')  # Redirect to a success page after approval



def e_voucher(request):
    # Fetch only approved vouchers
    approved_vouchers = CashVoucher.objects.filter(status='approved')
    
    if request.method == 'POST':
        # Handle form submission
        voucher_number = request.POST.get('voucher_number')
        # Use the selected voucher number for processing
        # ...

    return render(request, 'xp/e_voucher_page.html', {
        'approved_vouchers': approved_vouchers,
    })







from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .models import CashVoucher
from datetime import datetime

# This function will generate the voucher number
def generate_voucher_number():
    # Get the current date
    current_date = datetime.now()
    
    # Format the year as 'yy' and month as 'mm'
    year_month = current_date.strftime('%y%m')  # 'yy' format for year and 'mm' for month
    
    # Define the voucher counter. In practice, this would come from your database, and you would
    # increment the counter for each new voucher in the same month.
    # For demonstration purposes, we are assuming the counter starts at 1.
    last_voucher = CashVoucher.objects.filter(voucher_number__startswith=f"XPCV{year_month}").order_by('voucher_number').last()

    if last_voucher:
        # Extract the number part from the last voucher and increment it
        last_number = int(last_voucher.voucher_number[-4:])
        next_number = last_number + 1
    else:
        # If no voucher exists for this month, start from 1
        next_number = 1


    counter_str = str(next_number).zfill(4)

    # Combine everything into the final voucher number
    voucher_number = f"XPCV{year_month}{counter_str}"
    
    return voucher_number


# from django.shortcuts import render, redirect
# from .models import Voucher
# from .forms import VoucherForm

# def voucher_view(request):
#     if request.method == 'POST':
#         form = VoucherForm(request.POST, request.FILES)
#         if form.is_valid():
#             form.save()
#             return redirect('voucher_view')  # Redirect to the same page to display updated data
#     else:
#         form = VoucherForm()

#     # Fetch all vouchers (latest first)
#     vouchers = Voucher.objects.all().order_by('-voucher_number')
#     return render(request, 'xp/voucher_page.html', {'form': form, 'vouchers': vouchers})


from django.shortcuts import render
from .models import Expense

def e_voucher_page(request):
    """
    View to display all submitted expenses (e-vouchers).
    """
    # Fetch all Expense records, ordered by the most recent submission
    expenses = Expense.objects.all().order_by('-id')  # Assuming you have a 'created_at' timestamp field

    context = {
        'expenses': expenses,
    }

    return render(request, 'xp/e_voucher_page.html', context)


from django.db.models import Q
from django.core.paginator import Paginator
from django.shortcuts import render, redirect
from django.db.models import ForeignKey

def expense_form(request):
    # Retrieve search query from request
    search_query = request.GET.get('search', '')

    # Base queryset for all expenses
    expenses = Expense.objects.all()

    if search_query:
        # Initialize a Q object for global search
        global_filter = Q()

        # Loop through all fields of the Expense model
        for field in Expense._meta.fields:
            field_name = field.name

            if isinstance(field, ForeignKey):
                # Handle ForeignKey fields (searching related fields)
                related_model = field.related_model
                related_fields = [f.name for f in related_model._meta.fields if f.name != 'id']
                for related_field in related_fields:
                    global_filter |= Q(**{f"{field_name}__{related_field}__icontains": search_query})

            elif field.choices:  # Handle choice fields
                for value, display in dict(field.choices).items():
                    if search_query.lower() in display.lower():
                        global_filter |= Q(**{f"{field_name}": value})

            else:
                # Handle regular fields with icontains
                global_filter |= Q(**{f"{field_name}__icontains": search_query})

        # Filter expenses using the global_filter
        expenses = expenses.filter(global_filter)

    # Apply ordering and pagination
    expenses = expenses.order_by('-id')
    paginator = Paginator(expenses, 10)  # Show 10 records per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Add sequence numbers for the paginated records
    start_sequence_number = (page_obj.number - 1) * paginator.per_page + 1
    for index, expense in enumerate(page_obj, start=start_sequence_number):
        expense.sequence_number = index

    # Prepare the form for adding new expenses
    form = ExpenseForm()

    # Render the template with search results and form
    return render(request, 'xp/item_form.html', {
        'form': form,
        'page_obj': page_obj,
        'search_query': search_query,
    })
    

def expense_details(request, expense_id):
    """
    View to show the details of a specific expense.
    """
    expense = get_object_or_404(Expense, id=expense_id)
    context = {
        'expense': expense
    }
    return render(request, 'xp/expense_details.html', context)


# from django.shortcuts import render
# # from .forms import ItemForm
# from .models import Voucher

# def new_e_voucher(request):
#     # Handle form submission
#     if request.method == "POST":
#         form = ExpenseForm(request.POST, request.FILES)
#         if form.is_valid():
#             form.save()  # Save the form data to the database
#     else:
#         form = ExpenseForm()

#     # Retrieve all vouchers after form submission (including newly saved ones)
#     vouchers = Voucher.objects.all()  # Fetch all voucher data from the database

#     # Pass the form and vouchers data to the template
#     return render(request, 'xp/item_form.html', {'form': form, 'vouchers': vouchers})




from django.shortcuts import render, redirect
from .forms import ExpenseForm  # Or use ItemForm, if it's the form you're referring to
from .models import Expense

def expense_home(request):
    # Retrieve all Expense records
    expenses = Expense.objects.all()

    # If the form is submitted, handle it here
    if request.method == "POST":
        form = ExpenseForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()  # Save the new expense record
            return redirect('expense_home')  # Redirect back to this page after saving
    else:
        form = ExpenseForm()

    # Render the page with the form and all expenses
    return render(request, 'xp/expense_home.html', {
        'form': form,
        'expenses': expenses
    })

# views.py
from django.shortcuts import render, redirect
from django.contrib import messages
from django.utils import timezone
from .models import CashVoucher
from datetime import date

def new_cash_voucher_form(request):
    current_date = date.today()

    if request.method == 'POST':
        voucher_number = request.POST.get('voucher_number')
        cvamount = request.POST.get('cvamount')
        paid_to = request.POST.get('paid_to')
        item_name = request.POST.get('item_name')
        transaction_date = request.POST.get('transaction_date') 
        proof_photo = request.FILES.get('proof_photo')


        # Validate required fields
        if not cvamount or not paid_to or not item_name:

            return render(request, 'xp/new_cash_voucher_form.html', {'current_date': current_date },'voucher_number')

        try:
            # Create a new CashVoucher instance
            new_voucher = CashVoucher.objects.create(
                created_by=request.user,
                voucher_number=voucher_number,
                amount=cvamount,
                paid_to=paid_to,
                item_name=item_name,
                transaction_date=transaction_date,
		proof_photo=proof_photo,

            )


            return redirect('cash_voucher')  # Redirect back to the form page

        except Exception as e:
            messages.error(request, f"Error: {str(e)}")
            return redirect('new_cash_voucher_form')

    # GET request: fetch all existing vouchers
    cash_vouchers = CashVoucher.objects.all()
    cv_voucher_number =generate_voucher_number()
    context = {
        'current_date': current_date,
        'cash_vouchers': cash_vouchers,  # Pass the list of vouchers
        'cv_voucher_number':cv_voucher_number,
    }

    return render(request, 'xp/new_cash_voucher_form.html', context)


# views.py
from .models import Expense, CashVoucher, User, BorrowedAmount, AdvanceAssignment, AdvanceGroup, ProofPhoto, Conveyance
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from datetime import date, datetime
from decimal import Decimal, InvalidOperation


def edit_item(request, item_id):
    expense = get_object_or_404(Expense, id=item_id)
    current_date = date.today()
    selected_user = expense.created_by

    user_groups = AdvanceAssignment.objects.filter(
    user=request.user,
    group__is_closed=False
).select_related('group') 
    is_draft = expense.is_draft
    total_amount = expense.amount or 0
    transaction_category = expense.transaction_category or ""
    conveyance_options = Conveyance.objects.all()

    if request.method == "POST":
        data = request.POST
        files = request.FILES

        # --- Extract fields (same as new_item_form) ---
        item_type = data.get("item_type", expense.item_type)
        item_names = data.getlist("item_name", [])
        item_name = next((name for name in item_names if name.strip()), expense.item_name)
        transaction_option = data.get("transaction_option", expense.transaction_option).strip().lower()
        transaction_category = data.get("transaction_category", expense.transaction_category)
        kilometers = data.get("kilometers")
        vehicle_type = data.get("vehicle_type", transaction_option)
        internal_option = data.get("internal_option", "")
        external_type = data.get("external_type")
        total_amount = data.get("amount", expense.amount)
        payment_mode = data.get("payment_mode", expense.payment_mode)
        voucher_number = data.get("voucher_number", expense.voucher_number)
        evoucher_number = data.get("e_voucher_number", expense.evoucher_number)
        transaction_dates = data.getlist("transaction_date", [])
        transaction_date = next(
            (datetime.strptime(date.strip(), "%Y-%m-%d").date() for date in transaction_dates if date.strip()),
            expense.transaction_date,
        )

        # --- Handle bill/GST OCR override ---
        bill_photo = files.get("bill_photo")
        gst_photo = files.get("gst_photo")
        proof_files = files.getlist("proof_photos")


        # --- Conveyance calculation ---
        if transaction_option in ["2_wheeler", "4_wheeler"] and kilometers:
            try:
                kilometers = Decimal(kilometers)
                conveyance = Conveyance.objects.filter(vehicle_type=vehicle_type).first()
                if conveyance:
                    total_amount = kilometers * conveyance.price_per_km
                else:
                    messages.error(request, f"No conveyance rate found for {vehicle_type}.")
                    total_amount = 0
            except InvalidOperation:
                messages.error(request, "Invalid kilometers value.")
                total_amount = 0

        # --- Draft handling ---
        draft_status = data.get("draft_status")

        if draft_status == "true":
            is_draft = True
        else:
            is_draft = False   # 🔥 force exit draft on normal submit

        # --- AdvanceGroup deduction if applicable ---
        selected_group_id = data.get("selected_group_id")
        if item_type == "advances" and selected_group_id:
            try:
                group = AdvanceGroup.objects.get(id=selected_group_id)
                prev_amount = expense.amount or 0
                # Adjust balance: refund old amount, deduct new
                new_balance = group.remaining_balance + prev_amount - Decimal(total_amount)
                if new_balance < 0:
                    messages.error(request, "Insufficient group balance.")
                    return redirect("edit_item", item_id=item_id)
                group.used_advance = group.used_advance - prev_amount + Decimal(total_amount)
                group.save()
            except AdvanceGroup.DoesNotExist:
                messages.error(request, "Selected group not found.")
                return redirect("edit_item", item_id=item_id)

        # --- Determine approval and status ---
        bill_or_gst_present = bool(bill_photo or gst_photo)
        proof_present = bool(proof_files)
        needs_approval = True
        if transaction_option == "voucher":
            needs_approval = True
        elif item_type == "conveyance":
            needs_approval = False
        elif bill_or_gst_present and proof_present:
            needs_approval = False

        status = "paid" if item_type == "advances" and not needs_approval else "pending"

        # --- Update expense ---
        expense.item_type = item_type
        expense.item_name = item_name
        expense.transaction_option = transaction_option
        expense.transaction_category = transaction_category
        expense.internal_option = internal_option  # always a string
        expense.external_type = external_type if transaction_category == "external" else None
        expense.payment_mode = payment_mode
        expense.voucher_number = voucher_number if transaction_option == "voucher" else None
        expense.evoucher_number = evoucher_number
        expense.transaction_date = transaction_date
        expense.amount = total_amount
        expense.is_draft = is_draft
        expense.advance_group = group if item_type == "advances" else None
        expense.status = status
        expense.is_rejected = False
        expense.reject_reason = ""
        # ✅ Add this new block
        if transaction_option in ["2_wheeler", "4_wheeler"]:
            conveyance = Conveyance.objects.filter(vehicle_type=vehicle_type).first()
            expense.conveyance = conveyance
            if kilometers:
                expense.km = Decimal(kilometers)
        else:
            expense.conveyance = None
            expense.km = None

        if bill_photo:
            expense.bill_photo = bill_photo
        if gst_photo:
            expense.gst_photo = gst_photo

        expense.save()

        # --- Update proof photos ---
        if proof_files:
            ProofPhoto.objects.filter(expense=expense).delete()
            for file in proof_files:
                ProofPhoto.objects.create(expense=expense, file=file)

        # --- Handle borrowed amounts ---
        borrowed_amounts = data.getlist("borrowed_amounts[]")
        borrowed_froms = data.getlist("borrowed_froms[]")
        BorrowedAmount.objects.filter(expense=expense).delete()
        remaining_amount = Decimal(total_amount)
        for b_amt, b_from in zip(borrowed_amounts, borrowed_froms):
            if b_amt and b_from:
                b_amt_decimal = Decimal(b_amt)
                if b_amt_decimal <= remaining_amount:
                    BorrowedAmount.objects.create(expense=expense, amount=b_amt_decimal, borrowed_from_id=b_from)
                    # replicate for borrowed users
                    Expense.objects.create(
                        created_by=User.objects.get(id=b_from),
                        item_type=item_type,
                        item_name=f"Borrowed - {item_name}",
                        transaction_option=transaction_option,
                        transaction_category="internal",
                        internal_option=request.user,
                        amount=b_amt_decimal,
                        payment_mode=payment_mode,
                        voucher_number=voucher_number if transaction_option == "voucher" else None,
                        evoucher_number=evoucher_number,
                        transaction_date=transaction_date,
                    )
                    remaining_amount -= b_amt_decimal
        expense.amount = remaining_amount
        expense.save()

        # --- Cash Voucher / Tips handling (optional) ---
        if "tips_checkbox" in data and data.get("tips_checkbox") == "on":
            cv_amount = data.get("cvamount", "").strip()
            paid_to = data.get("paid_to", "").strip()
            cv_item_name = data.get("item_name", "").strip()
            cv_transaction_date = data.get("transaction_date", "").strip()
            cv_voucher_number = data.get("tip_voucher_number", "")
            tip_proof_photo = files.get("proof_photo")

            if cv_amount and paid_to and cv_item_name and cv_transaction_date:
                try:
                    cv_amount = Decimal(cv_amount)
                    CashVoucher.objects.create(
                        created_by=request.user,
                        amount=cv_amount,
                        paid_to=paid_to,
                        item_name=cv_item_name,
                        transaction_date=cv_transaction_date,
                        voucher_number=cv_voucher_number,
                        status="pending",
                        expense=expense,
                        proof_photo=tip_proof_photo,
                    )
                except InvalidOperation:
                    messages.error(request, "Invalid Cash Voucher amount.")

        messages.success(request, "Expense updated successfully!")

        # --- Redirect ---
        if is_draft:
            return redirect("draft_vouchers")
        return redirect("item_form")

    # --- GET request: pre-fill form ---
    users = User.objects.all()
    used_voucher_numbers = list(Expense.objects.values_list("voucher_number", flat=True).exclude(voucher_number__isnull=True).distinct())
    approved_vouchers = CashVoucher.objects.filter(created_by=selected_user, status="approved").exclude(voucher_number__in=used_voucher_numbers)

    submitted_data = None
    cv_voucher_number = expense.voucher_number or ""
    remaining_balance = None
    try:
        assignments = AdvanceAssignment.objects.filter(user=request.user).select_related("group")
        group = assignments.first().group if assignments.exists() else None
        remaining_balance = group.remaining_balance if group else None
    except AdvanceAssignment.DoesNotExist:
        remaining_balance = None
    borrowed_info = expense.borrowed_amounts.select_related("borrowed_from").all()
    selected_group = expense.advance_group
    # --- Normalize internal_option for template display ---
    internal_option_value = ""

    if hasattr(expense.internal_option, "first_name"):
        # Case 1: stored as a User object
        internal_option_value = expense.internal_option.first_name

    elif isinstance(expense.internal_option, str):
        # Case 2: stored as plain text or comma-separated
        parts = [p.strip() for p in expense.internal_option.split(",") if p.strip()]
        internal_option_value = parts if len(parts) > 1 else parts[0] if parts else ""

    elif isinstance(expense.internal_option, list):
        # Case 3: already a list
        internal_option_value = expense.internal_option

    print("DEBUG internal_option:", expense.internal_option)
    print("DEBUG type:", type(expense.internal_option))
    print("WORKING")
    context = {
        "expense": expense,
        "approved_vouchers": approved_vouchers,
        "current_date": current_date,
        "selected_user": selected_user,
        "users": users,
        "conveyance_options": conveyance_options,
        "submitted_data": submitted_data,
        "cv_voucher_number": cv_voucher_number,
        "remaining_balance": remaining_balance,
        "user_groups": user_groups,
        "borrowed_info": borrowed_info,
        "selected_group": selected_group,
        "internal_option_value": internal_option_value,
    }
    return render(request, "xp/edit_item.html", context)

from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from .models import CashVoucher  # Replace with your model name

def get_voucher_data(request):
    if request.method == "GET" and request.GET.get("id"):
        voucher_id = request.GET.get("id")
        voucher = get_object_or_404(CashVoucher, id=voucher_id)  # Fetch the voucher
        # Prepare data to return as JSON (voucher_id is not included)
        data = {
            "voucher_number": voucher.voucher_number,
            "amount": voucher.amount,
            "paid_to": voucher.paid_to,
            "item_name": voucher.item_name,
            "transaction_date": voucher.transaction_date,
        }
        return JsonResponse(data)  # Return the data in JSON format

from django.shortcuts import render, get_object_or_404, redirect
from .models import CashVoucher
from .forms import CashVoucherForm

def edit_cash_voucher_form(request, voucher_id):
    voucher = get_object_or_404(CashVoucher, id=voucher_id)

    if request.method == 'POST':
        form = CashVoucherForm(request.POST, request.FILES, instance=voucher, user=request.user)
        if form.is_valid():
            edited_voucher = form.save(commit=False)

            # ✅ If previously rejected, update status to 'pending'
            if edited_voucher.status == 'rejected':
                edited_voucher.status = 'pending'
                edited_voucher.rejection_reason = ""  # Optional: clear rejection reason

            edited_voucher.save()
            form.save_m2m()  # In case your form has many-to-many fields

            return redirect('item_form')  # Or wherever you want to redirect
        else:
            print("Form errors:", form.errors)
    else:
        form = CashVoucherForm(instance=voucher, user=request.user)

    return render(request, 'xp/edit_cash_voucher.html', {'form': form, 'voucher': voucher})



def save_expense(request):
    if request.method == 'POST':
        # Assuming you are using a form to save the data
        form = ExpenseForm(request.POST, request.FILES)
        if form.is_valid():
            # Capture internal or external based on the category
            transaction_category = form.cleaned_data.get('transaction_category')

            # If internal, save the internal_option
            if transaction_category == 'internal':
                internal_option = form.cleaned_data.get('internal_option')
                expense = form.save(commit=False)
                expense.internal_option = internal_option
                expense.save()

            # If external, save the external_type
            elif transaction_category == 'external':
                external_type = form.cleaned_data.get('external_type')
                expense = form.save(commit=False)
                expense.external_type = external_type
                expense.save()

            # If no specific category, just save the form
            else:
                expense = form.save()

            return redirect('success_url')  # Redirect to success page

    else:
        form = ExpenseForm()

    return render(request, 'xp/expense_form.html', {'form': form})


from django.shortcuts import render
from django.db.models import Q
from .models import Expense  # Assuming you're searching through the 'Expense' model

def search_view(request):
    search_query = request.GET.get('search', '')
    expenses = Expense.objects.all()

    if search_query:
        # Perform a case-insensitive search across multiple fields
        expenses = expenses.filter(
            Q(item_name__icontains=search_query) |
            Q(item_type__icontains=search_query) |
            Q(payment_category__icontains=search_query) |
            Q(transaction_category__icontains=search_query) |
            Q(transaction_details__icontains=search_query)
        )

    return render(request, 'xp/item_form.html', {'expenses': expenses, 'search_query': search_query})


from django.shortcuts import get_object_or_404
from django.http import JsonResponse
from .models import Expense

def update_item(request):
    if request.method == 'POST':
        # Get the expense ID from the POST data
        expense_id = request.POST.get('id')
        if not expense_id:
            return JsonResponse({'error': 'ID is required'}, status=400)

        # Get the Expense object or return a 404 if it doesn't exist
        expense = get_object_or_404(Expense, id=expense_id)

        # Update regular fields
        expense.item_type = request.POST.get('item_type')
        expense.item_name = request.POST.get('item_name')
        expense.payment_category = request.POST.get('transaction_option')
        expense.transaction_category = request.POST.get('transaction_category')
        expense.amount = request.POST.get('amount')
        expense.transaction_date = request.POST.get('transaction_date')

        # Handle file uploads
        if 'bill_photo' in request.FILES:
            expense.bill_photo = request.FILES['bill_photo']
        if 'gst_photo' in request.FILES:
            expense.gst_photo = request.FILES['gst_photo']
        if 'voucher_number' in request.POST:
            expense.voucher_number = request.POST.get('voucher_number')

        # If there are transaction-related fields, set them accordingly
        if expense.transaction_category == 'internal':
            expense.internal_option = request.POST.get('internal_option', '')
        elif expense.transaction_category == 'external':
            expense.external_detail = request.POST.get('external_type', '')

        # Save the updated object
        expense.save()

        return JsonResponse({'message': 'Item updated successfully'})

    return JsonResponse({'error': 'Invalid request method'}, status=405)


    

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .models import Expense
from datetime import datetime

def generate_evoucher_number():
    current_date = datetime.now()
    year_month = current_date.strftime('%y%m')  # e.g., '2412' for Dec 2024

    last_voucher = Expense.objects.filter(
        evoucher_number__startswith=f"XPEV{year_month}"
    ).order_by('-evoucher_number').first()

    if last_voucher:
        last_number = int(last_voucher.evoucher_number[-4:])
        next_number = last_number + 1
    else:
        next_number = 1

    counter_str = str(next_number).zfill(4)
    evoucher_number = f"XPEV{year_month}{counter_str}"
    return evoucher_number

from django.utils.dateparse import parse_date
from datetime import datetime, timedelta

def get_filtered_expenses(request):
    print("GOT IT")
    expenses = Expense.objects.all()
    filter_field = 'transaction_date'

    # Reads query params from modal or quick links
    from_date = request.GET.get('transaction_date_from') or request.GET.get('from')
    print("from date",from_date)
    to_date   = request.GET.get('transaction_date_to') or request.GET.get('to')
    print("To date",to_date)
    days      = request.GET.get('days')
    item_type = request.GET.get('item_type')

    if item_type:
        expenses = expenses.filter(item_type=item_type)

    if days:
        try:
            days = int(days)
            filter_date = datetime.now().date() - timedelta(days=days)
            expenses = expenses.filter(**{f"{filter_field}__gte": filter_date})
        except Exception:
            pass

    elif from_date and to_date:
        from_dt = parse_date(from_date)
        to_dt = parse_date(to_date)
        if from_dt and to_dt:
            expenses = expenses.filter(**{f"{filter_field}__range": (from_dt, to_dt)})

    return expenses







# def export_vouchers(request):
#     """
#     View for exporting vouchers to Excel, PDF, or CSV.
#     """
#     # Add logic for exporting data based on request parameters (e.g., export format)
#     export_format = request.GET.get('format', 'xlsx')  # Can be 'xlsx', 'pdf', or 'csv'
    
#     if export_format == 'xlsx':
#         return export_to_xlsx(request)
#     elif export_format == 'pdf':
#         return export_to_pdf(request)
#     elif export_format == 'csv':
#         return export_to_csv(request)
#     else:
#         return JsonResponse({'error': 'Invalid export format'})

from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_GET
import xlsxwriter
from io import BytesIO
import csv
from xhtml2pdf import pisa

from .models import Expense  # adjust import as needed

# def get_filtered_expenses(request):
#     # Replace with your filtering logic!
#     return Expense.objects.all()

@require_GET
def export_vouchers(request):
    """
    Master export dispatcher for vouchers/expenses.
    """
    export_format = request.GET.get('format', 'xlsx')
    if export_format == 'xlsx':
        return export_to_xlsx(request)
    elif export_format == 'pdf':
        return export_to_pdf(request)
    elif export_format == 'csv':
        return export_to_csv(request)
    else:
        return JsonResponse({'error': 'Invalid export format'})



def export_to_xlsx(request):
    import xlsxwriter
    from io import BytesIO
    from django.http import HttpResponse
    from decimal import Decimal

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=vouchers.xlsx'

    output = BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet('Vouchers')

    # --- Formats ---
    header_format = workbook.add_format({
        'bold': True, 'border': 1, 'align': 'center', 'valign': 'vcenter',
        'fg_color': '#0070C0', 'font_color': 'white'
    })
    blue_format = workbook.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter', 'fg_color': '#B4C6E7'})
    white_format = workbook.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter', 'fg_color': '#D9E1F2'})
    link_blue_format = workbook.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter', 'color': 'blue', 'underline': 1, 'fg_color': '#B4C6E7'})
    link_white_format = workbook.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter', 'color': 'blue', 'underline': 1, 'fg_color': '#D9E1F2'})
    date_blue_format = workbook.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter', 'fg_color': '#B4C6E7', 'num_format': 'd-mmm-yyyy'})
    date_white_format = workbook.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter', 'fg_color': '#D9E1F2', 'num_format': 'd-mmm-yyyy'})

    # --- Headers ---
    headers = [
        "Serial No", "Date", "Item Name", "Item Type", "Payment Category",
        "Paid By", "Paid To", "Transaction Category",
        "Mode of Payment", "Amount", "Bill/GST Download", "Proof Download"
    ]
    worksheet.write_row(0, 0, headers, header_format)
    user_id = request.GET.get("created_by")

    # Fetch expenses
    expenses = get_filtered_expenses(request).filter(
        is_draft=False, delete_requested=False
    ).order_by("transaction_date", "id")
    site_url = request.build_absolute_uri('/')[:-1]
    if user_id:
        expenses = expenses.filter(created_by_id=user_id)


    # Filter expenses: only include fully attached or approved
    filtered_expenses = []
    for exp in expenses:
        if exp.item_type.lower() == "conveyance":
            filtered_expenses.append(exp)
            continue

        has_bill_gst = (exp.bill_photo or exp.gst_photo)
        has_proof = exp.proof_photo or exp.proof_photos.exists()

        if (has_bill_gst and has_proof) or getattr(exp, 'is_approved', False):
            filtered_expenses.append(exp)

    internal_total_amount = Decimal("0.00")
    internal_count = 0
    external_total_amount = Decimal("0.00")
    external_count = 0

    # --- Write Expenses ---
    serial = 1
    row_idx = 1
    for expense in filtered_expenses:
        date_value = expense.transaction_date
        item_name = expense.item_name or ""
        item_type = expense.item_type or ""
        payment_category = dict(expense.PAYMENT_OPTIONS).get(
            expense.transaction_option, expense.transaction_option or ""
        )
        paid_by = expense.created_by.first_name or expense.created_by.username.split("_")[0] if expense.created_by else ""

        # Paid To
        if expense.transaction_category == "internal":
            paid_to = expense.internal_option or ""
            if expense.advance_group and expense.advance_group.leader:
                leader_name = expense.advance_group.leader.get_full_name() or expense.advance_group.leader.username
                paid_to = f"{paid_to} {leader_name}"
        elif expense.transaction_category == "external":
            paid_to = expense.external_type or ""
        else:
            paid_to = ""

        t_category = expense.transaction_category or ""
        mode_of_payment = expense.payment_mode or ""
        amount = float(expense.amount or 0)

        # Bill/GST link
        bill_link = None
        if expense.item_type.lower() != "conveyance":
            if expense.bill_photo and hasattr(expense.bill_photo, "url"):
                bill_link = (site_url + expense.bill_photo.url, "Download Bill/GST")
            elif expense.gst_photo and hasattr(expense.gst_photo, "url"):
                bill_link = (site_url + expense.gst_photo.url, "Download GST")

        # Proof links
        proof_links = []
        if expense.item_type.lower() != "conveyance":
            proofs = list(expense.proof_photos.all())
            if not proofs and expense.proof_photo and hasattr(expense.proof_photo, "url"):
                proofs = [expense]
            for proof in proofs:
                if hasattr(proof, "file") and proof.file:
                    proof_links.append((site_url + proof.file.url, "Download Proof"))
                elif hasattr(proof, "proof_photo") and proof.proof_photo:
                    proof_links.append((site_url + proof.proof_photo.url, "Download Proof"))

        if expense.transaction_category == 'internal':
            internal_total_amount += Decimal(expense.amount or 0)
            internal_count += 1
        elif expense.transaction_category == 'external':
            external_total_amount += Decimal(expense.amount or 0)
            external_count += 1

        # Alternate row color selection
        if row_idx % 2 == 0:
            row_fmt = white_format
            link_fmt = link_white_format
            date_fmt = date_white_format
        else:
            row_fmt = blue_format
            link_fmt = link_blue_format
            date_fmt = date_blue_format

        row_data = [
            serial, date_value, item_name, item_type, payment_category,
            paid_by, paid_to, t_category,
            mode_of_payment, amount, "", ""
        ]

        for col_idx, value in enumerate(row_data):
            if col_idx == 1 and date_value:
                worksheet.write_datetime(row_idx, col_idx, date_value, date_fmt)
            else:
                worksheet.write(row_idx, col_idx, value, row_fmt)

        if bill_link:
            worksheet.write_url(row_idx, 10, bill_link[0], link_fmt, string=bill_link[1])

        if proof_links:
            if len(proof_links) > 1:
                for col_fill in range(11):
                    worksheet.merge_range(
                        row_idx, col_fill, row_idx + len(proof_links) - 1, col_fill,
                        row_data[col_fill], row_fmt
                    )
            for i, (url, label) in enumerate(proof_links):
                worksheet.write_url(row_idx + i, 11, url, link_fmt, string=label)
            row_idx += len(proof_links) - 1

        serial += 1
        row_idx += 1

    # Auto-adjust column width
    for i, header in enumerate(headers):
        worksheet.set_column(i, i, max(len(header)+2, 20))

    # --- Summary Table ---
    summary_start = row_idx + 2
    worksheet.write(summary_start, 0, "Summary", header_format)
    worksheet.write_row(summary_start + 1, 0, ["Type", "Total Transactions", "Total Amount"], header_format)

    summary_data = [
        ["Internal", internal_count, float(internal_total_amount)],
        ["External", external_count, float(external_total_amount)]
    ]
    for i, row in enumerate(summary_data):
        for j, val in enumerate(row):
            worksheet.write(summary_start + 2 + i, j, val, blue_format)

    workbook.close()
    response.write(output.getvalue())
    output.close()
    return response

def export_to_pdf(request):
    from django.http import HttpResponse
    from xhtml2pdf import pisa
    from io import BytesIO

    vouchers = get_filtered_expenses(request)
    site_url = request.build_absolute_uri('/')[:-1]
    html_content = """
    <html>
    <head><style>
    table {border-collapse: collapse; width: 100%;}
    th, td {border: 1px solid #ddd; padding: 8px;}
    a {color: blue;}
    </style></head>
    <body>
    <h2>E-Voucher Export</h2>
    <table>
        <thead>
            <tr>
                <th>SLNO</th>
                <th>Transaction Date</th>
                <th>Item Name</th>
                <th>Item Type</th>
                <th>Payment Category</th>
                <th>Paid By</th>
                <th>Paid To</th>
                <th>Model of Payment</th>
                <th>Amount</th>
                <th>Bill/GST</th>
                <th>Proof</th>
            </tr>
        </thead>
        <tbody>
    """
    for idx, v in enumerate(vouchers, start=1):
        bill_url = site_url + v.bill_photo.url if v.bill_photo and hasattr(v.bill_photo, "url") else ""
        gst_url  = site_url + v.gst_photo.url  if v.gst_photo  and hasattr(v.gst_photo,  "url") else ""
        bill_gst_cell = ""
        if bill_url and gst_url:
            bill_gst_cell = f'<a href="{bill_url}">Download Bill</a><br><a href="{gst_url}">Download GST</a>'
        elif bill_url:
            bill_gst_cell = f'<a href="{bill_url}">Download Bill</a>'
        elif gst_url:
            bill_gst_cell = f'<a href="{gst_url}">Download GST</a>'

        # All proofs
        proof_links = []
        for proof in v.proof_photos.all():
            if proof.file and hasattr(proof.file, "url"):
                url = site_url + proof.file.url
                proof_links.append(f'<a href="{url}">Download Proof</a>')
        proof_cell = "<br>".join(proof_links)

        html_content += f"""
        <tr>
            <td>{idx}</td>
            <td>{v.transaction_date.strftime('%Y-%m-%d') if v.transaction_date else ''}</td>
            <td>{v.item_name or ''}</td>
            <td>{v.item_type or ''}</td>
            <td>{dict(v.PAYMENT_OPTIONS).get(v.transaction_option, v.transaction_option or "")}</td>
            <td>{str(v.created_by) if v.created_by else ''}</td>
            <td>{v.external_type or ''}</td>
            <td>{v.payment_mode or ''}</td>
            <td>{float(v.amount) if v.amount is not None else ''}</td>
            <td>{bill_gst_cell}</td>
            <td>{proof_cell}</td>
        </tr>
        """

    html_content += "</tbody></table></body></html>"

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="vouchers.pdf"'
    pdf_buffer = BytesIO()
    pisa.CreatePDF(html_content, dest=pdf_buffer)
    pdf_buffer.seek(0)
    response.write(pdf_buffer.read())
    return response


def export_to_csv(request):
    import csv
    from django.http import HttpResponse

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=vouchers.csv'

    writer = csv.writer(response)
    writer.writerow([
        "SLNO",
        "Transaction Date",
        "Item Name",
        "Item Type",
        "Payment Category",
        "Paid By",
        "Paid To",
        "Model of Payment",
        "Amount",
        "Bill/GST",    # <--- one column
        "Proof",       # <--- one column, all proof files comma-separated
    ])
    expenses = get_filtered_expenses(request)
    site_url = request.build_absolute_uri('/')[:-1]

    for idx, expense in enumerate(expenses, start=1):
        payment_category = dict(expense.PAYMENT_OPTIONS).get(expense.transaction_option, expense.transaction_option or "")
        transaction_date = expense.transaction_date.strftime('%Y-%m-%d') if expense.transaction_date else ""
        paid_by = str(expense.created_by) if expense.created_by else ""
        paid_to = expense.external_type or ""
        amount = float(expense.amount) if expense.amount is not None else ""
        payment_mode = expense.payment_mode or ""

        # Bill/GST
        bill_url = site_url + expense.bill_photo.url if expense.bill_photo and hasattr(expense.bill_photo, "url") else ""
        gst_url  = site_url + expense.gst_photo.url  if expense.gst_photo  and hasattr(expense.gst_photo,  "url") else ""
        bill_gst_col = ""
        if bill_url and gst_url:
            bill_gst_col = ', '.join([bill_url, gst_url])
        elif bill_url:
            bill_gst_col = bill_url
        elif gst_url:
            bill_gst_col = gst_url
        # (Or, if you want ONLY Bill when both are present, use just bill_url)

        # All proofs
        proof_links = []
        for proof in expense.proof_photos.all():
            if proof.file and hasattr(proof.file, "url"):
                proof_links.append(site_url + proof.file.url)
        proof_col = ', '.join(proof_links)

        writer.writerow([
            idx,
            transaction_date,
            expense.item_name or "",
            expense.item_type or "",
            payment_category,
            paid_by,
            paid_to,
            payment_mode,
            amount,
            bill_gst_col,
            proof_col,
        ])
    return response




from django.http import FileResponse, Http404
import os
from django.conf import settings

def download_proof(request, file_name):
    """
    Serve the uploaded proof file as an attachment for download.
    """
    file_path = os.path.join(settings.MEDIA_ROOT, file_name)
    
    if not os.path.exists(file_path):
        raise Http404("File not found.")

    response = FileResponse(open(file_path, 'rb'), as_attachment=True)
    return response

from django.shortcuts import render
from .models import  CashVoucher

def search_vouchers(request):
    query = request.GET.get('q', '')
    vouchers = []
    cash_vouchers = []

    if query:
        # Search in Voucher and CashVoucher models
        vouchers = CashVoucher.objects.filter(
            voucher_number__icontains=query
        ) | CashVoucher.objects.filter(
            item_name__icontains=query
        )

        cash_vouchers = CashVoucher.objects.filter(
            voucher_number__icontains=query
        ) | CashVoucher.objects.filter(
            paid_to__icontains=query
        )

    context = {
        'vouchers': vouchers,
        'cash_vouchers': cash_vouchers,
        'query': query,
    }
    return render(request, 'xp/voucher_list.html', context)


import csv
from io import BytesIO
from django.http import HttpResponse
import xlsxwriter
from reportlab.pdfgen import canvas
from .models import CashVoucher  # Replace with your actual model


def export_cash_vouchers(request):
    format = request.GET.get('format', 'csv')
    vouchers = CashVoucher.objects.all()  # Adjust queryset as needed for filtering
    
    if format == 'csv':
        return export_cash_vouchers_csv(vouchers)
    elif format == 'xlsx':
        return export_cash_vouchers_xlsx(vouchers)
    elif format == 'pdf':
        return export_cash_vouchers_pdf(vouchers)
    else:
        return HttpResponse("Invalid format", status=400)


def export_cash_vouchers_csv(vouchers):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="cash_vouchers.csv"'
    writer = csv.writer(response)
    
    # Write the header
    writer.writerow(['Voucher Number', 'Date', 'Paid To', 'Item Name', 'Amount', 'Transaction Date', 'Status'])
    
    # Write the data rows
    for voucher in vouchers:
        writer.writerow([
            voucher.voucher_number,
            voucher.date,
            voucher.paid_to,
            voucher.item_name,
            voucher.amount,
            voucher.transaction_date,
            voucher.status,
        ])
    
    return response


from django.http import HttpResponse
import openpyxl  # or any library you're using to generate the XLSX file

def export_cash_vouchers_xlsx(vouchers):
    # Generate your XLSX file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Voucher Number', 'Item Name', 'Amount'])  # Example headers
    
    # Add rows for each voucher
    for voucher in vouchers:
        ws.append([voucher.voucher_number,
            voucher.date,
            voucher.paid_to,
            voucher.item_name,
            voucher.amount,
            voucher.transaction_date,
            voucher.status,])

    # Create an HTTP response with the XLSX content
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Cash_Vouchers.xlsx'
    wb.save(response)
    
    return response


from io import BytesIO
from xhtml2pdf import pisa
from django.http import HttpResponse
from django.template.loader import render_to_string

def export_cash_vouchers_pdf(request):
    """
    Export vouchers to PDF using a styled HTML template.
    """
    # Fetch the vouchers you want to export
    vouchers = CashVoucher.objects.all()

    # Prepare the context for the HTML template
    context = {
        'vouchers': vouchers
    }

    # Render the HTML content using Django templates
    html_content = render_to_string('xp/export_cash_vouchers_template.html', context)

    # Create an HttpResponse object with PDF headers
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="CashVoucher.pdf"'

    # Convert the HTML content to PDF
    pdf_buffer = BytesIO()
    pisa_status = pisa.CreatePDF(html_content, dest=pdf_buffer)

    if pisa_status.err:
        return HttpResponse('Error generating PDF', status=500)

    # Write the PDF content to the response
    pdf_buffer.seek(0)
    response.write(pdf_buffer.read())
    return response
    
from django.shortcuts import render
from django.db.models import Sum
from django.db.models import F, Func
from .models import CashVoucher
from .models import Expense  # Assuming Expense model is in the `expenses` app

from django.db.models.functions import TruncMonth


from datetime import date, datetime, timedelta


from django.shortcuts import render
from django.db.models import Sum
from django.db.models import F, Func
from .models import CashVoucher
from .models import Expense  # Assuming Expense model is in the `expenses` app
from django.db import connection
from django.db.models import Sum, F
from django.db.models.functions import TruncMonth


from datetime import date, datetime, timedelta


from django.db.models import Sum, Count, Q
from django.db.models.functions import TruncMonth
from datetime import date, datetime

def vouchers_dashboard(request):
    user = request.user

    # Get date filters from request
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    # Default to the current month if no filters are provided
    if not from_date or not to_date:
        today = date.today()
        first_day = today.replace(day=1)
        from_date = first_day
        to_date = today

    # Convert to date objects if they are strings
    if isinstance(from_date, str):
        from_date = datetime.strptime(from_date, "%Y-%m-%d").date()
    if isinstance(to_date, str):
        to_date = datetime.strptime(to_date, "%Y-%m-%d").date()

    # Filter vouchers and expenses based on user role and date range
    if user.is_staff:
        vouchers = CashVoucher.objects.filter(date__range=(from_date, to_date))
        expenses = Expense.objects.filter(transaction_date__range=(from_date, to_date), is_draft=False)
    else:
        vouchers = CashVoucher.objects.filter(created_by=user, date__range=(from_date, to_date))
        expenses = Expense.objects.filter(created_by=user, transaction_date__range=(from_date, to_date), is_draft=False)

    # Calculate total counts for each status
    total_vouchers = vouchers.count()
    approved_vouchers = vouchers.filter(status='approved').count()
    rejected_vouchers = vouchers.filter(status='rejected').count()
    pending_vouchers_count = vouchers.filter(status='pending').count()
    cleared_vouchers_count = vouchers.filter(status='paid').count()

    # Calculate total claiming amount and approved amount
    total_claiming_amount = expenses.aggregate(Sum('amount'))['amount__sum'] or 0
    total_claiming_amount_cv = vouchers.aggregate(Sum('amount'))['amount__sum'] or 0
    total_approved_amount = vouchers.filter(status='approved').aggregate(Sum('amount'))['amount__sum'] or 0

    # Aggregate the total paid amount for expenses and vouchers
    if user.is_staff:
        paid_expenses = Expense.objects.filter(status='paid')
        paid_vouchers = CashVoucher.objects.filter(status='paid')
    else:
        paid_expenses = Expense.objects.filter(created_by=user, status='paid')
        paid_vouchers = CashVoucher.objects.filter(created_by=user, status='paid')

    # Aggregate the total paid amount
    total_paid_amount = paid_expenses.aggregate(Sum('amount'))['amount__sum'] or 0
    total_paid_amount_cv = paid_vouchers.aggregate(Sum('amount'))['amount__sum'] or 0

    # Total e-voucher count
    total_evouchers = expenses.count()

    # Difference between claiming and approved amounts
    claiming_approved_diff = total_claiming_amount - total_paid_amount

    # Calculate month-wise total claimed and approved amounts
    vouchers_by_month = vouchers.annotate(
        month=TruncMonth('date')
    ).values('month').annotate(
        total_claimed=Sum('amount'),
        total_paid=Sum('amount', filter=Q(status='paid')),
        total_evoucher_amount=Sum('amount', filter=Q(status='approved'))
    ).order_by('month')
    

    months_list = []
    claimed_amounts = []
    paid_amounts = []
    evoucher_amounts = []

    for voucher in vouchers_by_month:
        months_list.append(voucher['month'].strftime('%b %Y'))
        claimed_amounts.append(voucher['total_claimed'])
        paid_amounts.append(voucher['total_paid'])
        evoucher_amounts.append(voucher['total_evoucher_amount'])

    # Item type counts for chart rendering
    item_type_counts_and_amounts = (
    expenses.values('item_type')
    .annotate(
        count=Count('item_type'),
        total_amount=Sum('amount')  # Assuming 'amount' is the field where the amount is stored
    )
)

# Transform item type counts and amounts into a dictionary
    item_type_data = {
            entry['item_type']: {
                'count': entry['count'],
                'total_amount': entry['total_amount'] if entry['total_amount'] is not None else 0
            }
            for entry in item_type_counts_and_amounts
        }

# Ensure all item types are represented, even with a count of 0
    all_item_types = [
            'fright_transportation', 'fuel_oil', 'food_grocery', 'others', 'vendor',
            'reimbursement', 'site_services', 'advances', 'travelling', 'electricity_bill',
            'internet_bill', 'postage_telegram', 'machine_repair', 'conveyance',
            'stationary_printing', 'hospitality'
        ]

    item_type_data = {item_type: item_type_data.get(item_type, {'count': 0, 'total_amount': 0}) for item_type in all_item_types}

    # Context for template rendering
    context = {
        'from_date': from_date,
        'to_date': to_date,
        'total_vouchers': total_vouchers,
        'approved_vouchers': approved_vouchers,
        'rejected_vouchers': rejected_vouchers,
        'pending_vouchers_count': pending_vouchers_count,
        'total_claiming_amount': total_claiming_amount,
        'total_approved_amount': total_approved_amount,
        'total_evouchers': total_evouchers,
        'claiming_approved_diff': claiming_approved_diff,
        'months_list': months_list,
        'claimed_amounts': claimed_amounts,
        'paid_amounts': paid_amounts,
        'total_paid_amount': total_paid_amount,
        'cleared_vouchers_count': cleared_vouchers_count,
        'total_claiming_amount_cv': total_claiming_amount_cv,
        'total_paid_amount_cv': total_paid_amount_cv,
        'evoucher_amounts': evoucher_amounts,
        'item_type_data': item_type_data,  # Pass item type data for chart
    }

    return render(request, 'xp/dashboard.html', context)


from django.http import JsonResponse
from django.shortcuts import render
from .models import CashVoucher, User
from app.models import Zone, UserProfile
from datetime import timedelta, date
from django.db.models import Q

def get_users_by_zone(request):
    zone_id = request.GET.get('zone_id')
    users = []
    
    if zone_id:
        selected_zone = Zone.objects.get(id=zone_id)
        users_in_zone = UserProfile.objects.filter(zone=selected_zone).values(
            'user__id',
            'user__username',
            'user__first_name'
        )

        users = [
            {
                'id': user['user__id'],
                'username': user['user__username'],
                'first_name': user['user__first_name'] or ""   # include first name
            }
            for user in users_in_zone
        ]

    return JsonResponse({'users': users})

# from datetime import date, timedelta
# from django.shortcuts import render
# from .models import User, Zone, CashVoucher, UserProfile
#without leader balance subtraction
# def payable(request):
#     users = User.objects.all()
#     zones = Zone.objects.all()
#     selected_user = None
#     selected_zone = None
#     from_date = request.POST.get('from_date')
#     to_date = request.POST.get('to_date')
#     filtered_users = User.objects.none()
#     total_amount_to_be_paid = Decimal('0.00')
#     min_date = None
#     max_date = None
#     expenses = []  # empty by default

#     if request.method == "POST":
#         queryset = Expense.objects.all().select_related("advance_group")
#         user_id = request.POST.get("user_id")
#         zone_id = request.POST.get("zone_id")

#         # Zone filter
#         if zone_id:
#             selected_zone = get_object_or_404(Zone, id=zone_id)
#             users_in_zone = UserProfile.objects.filter(zone=selected_zone).values_list("user", flat=True)
#             filtered_users = User.objects.filter(id__in=users_in_zone)
#             queryset = queryset.filter(created_by__in=users_in_zone)

#         # User filter
#         if user_id:
#             selected_user = get_object_or_404(User, id=user_id)
#             queryset = queryset.filter(created_by=selected_user)

#         # Date filter
#         if from_date and to_date:
#             try:
#                 from_date_parsed = datetime.strptime(from_date, "%Y-%m-%d").date()
#                 to_date_parsed = datetime.strptime(to_date, "%Y-%m-%d").date()
#                 queryset = queryset.filter(transaction_date__gte=from_date_parsed,
#                                            transaction_date__lte=to_date_parsed)
#             except ValueError:
#                 pass

#         if queryset.exists():
#             min_date = queryset.earliest("transaction_date").transaction_date
#             max_date = queryset.latest("transaction_date").transaction_date

#         # Annotate proof counts
#         expenses = queryset.annotate(proof_count=Count("proof_photos")).order_by("-transaction_date", "-id")

#         # Track remaining per advance group
#         advance_group_remaining = {}
#         advance_due_expenses = []

#         for exp in expenses:
#             exp.overdue_amount = Decimal('0.00')
#             if exp.item_type != "advances" and exp.amount:
#                 total_amount_to_be_paid += exp.amount
#             if exp.item_type == "advances" and exp.advance_group:
#                 ag_id = exp.advance_group.id
#                 if ag_id not in advance_group_remaining:
#                     advance_group_remaining[ag_id] = exp.advance_group.total_advance
#                 advance_group_remaining[ag_id] -= exp.amount or Decimal("0.00")
#                 if advance_group_remaining[ag_id] < 0:
#                     exp.overdue_amount = abs(advance_group_remaining[ag_id])
#                     total_amount_to_be_paid += exp.overdue_amount
#                     advance_group_remaining[ag_id] = Decimal('0.00')
#                     advance_due_expenses.append(exp)
#     else:
#         # no POST → show nothing
#         advance_due_expenses = []

#     context = {
#         "users": users,
#         "zones": zones,
#         "expenses": expenses,
#         "advance_due_expenses": advance_due_expenses,
#         "selected_user": selected_user,
#         "selected_zone": selected_zone,
#         "from_date": from_date,
#         "to_date": to_date,
#         "filtered_users": filtered_users,
#         "total_amount_to_be_paid": total_amount_to_be_paid,
#         "min_date": min_date,
#         "max_date": max_date,
#     }

#     return render(request, "xp/payable.html", context)


from django.db.models import Q, Count, Sum
from datetime import datetime
from decimal import Decimal
from collections import defaultdict
from app.models import Zone, UserProfile   # ✅ Correct import

from django.db.models import Q, Count, Sum
from datetime import datetime
from decimal import Decimal
from collections import defaultdict
from app.models import Zone, UserProfile   # correct import

def payable(request):
    users = User.objects.all()
    zones = Zone.objects.all()
    expenses = Expense.objects.none()
    selected_user = None
    selected_zone = None
    filtered_users = User.objects.none()

    total_amount_to_be_paid = Decimal("0.00")
    final_amount_to_pay = Decimal("0.00")

    leader_sub_balance = Decimal("0.00")
    leader_overdue_total = Decimal("0.00")
    leader_group_balances = []

    from_date = request.POST.get("from_date")
    to_date = request.POST.get("to_date")

    queryset = Expense.objects.all()

    if request.method == "POST":
        user_id = request.POST.get("user_id")
        zone_id = request.POST.get("zone_id")

        queryset = Expense.objects.all()

        # ------------------ ZONE FILTER ---------------------
        if zone_id and UserProfile:
            selected_zone = Zone.objects.get(id=zone_id)
            users_in_zone = UserProfile.objects.filter(zone=selected_zone).values_list("user", flat=True)
            filtered_users = User.objects.filter(id__in=users_in_zone)
            queryset = queryset.filter(created_by__in=users_in_zone)

        # ------------------ USER FILTER ---------------------
        if user_id:
            selected_user = User.objects.get(id=user_id)
            queryset = queryset.filter(created_by=selected_user)

        # ------------------ DATE FILTER (Expenses only) -----
        date_filter_expenses = {}
        if from_date and to_date:
            try:
                f = datetime.strptime(from_date, "%Y-%m-%d").date()
                t = datetime.strptime(to_date, "%Y-%m-%d").date()
                date_filter_expenses = {"transaction_date__gte": f, "transaction_date__lte": t}
                queryset = queryset.filter(**date_filter_expenses)
            except:
                pass

        # ------------------ EXPENSE QUERY ---------------------
        expenses = (
            queryset
                .filter(is_allocation=False)  # 🔥 DO NOT include allocated vouchers
                .annotate(proof_count=Count("proof_photos"))
                .filter(
                    Q(item_type="conveyance")
                    | Q(is_approved=True)
                    | (
                        ~Q(item_type="conveyance")
                        & ~Q(transaction_option="voucher")
                        & (
                            (Q(bill_photo__isnull=False) & ~Q(bill_photo=""))
                            | (Q(gst_photo__isnull=False) & ~Q(gst_photo=""))
                        )
                        & (
                            (Q(proof_photo__isnull=False) & ~Q(proof_photo=""))
                            | Q(proof_count__gt=0)
                        )
                    )
                )
                .exclude(Q(status="paid") | Q(status="rejected") | Q(is_rejected=True))
                .filter(is_draft=False)
                .filter(Q(cash_vouchers__isnull=True) | Q(cash_vouchers__status="approved"))
        )

        # ------------------ OVERDUE CHECK ---------------------
        if selected_user:
            for exp in expenses:
                if exp.advance_group:
                    used_total = Expense.objects.filter(
                        advance_group=exp.advance_group,
                        is_allocation=False,
                        is_draft=False,
                        status__in=["pending", "approved", "paid"],
                        **date_filter_expenses
                    ).aggregate(total=Sum("amount"))["total"] or Decimal("0.00")

                    group_total = exp.advance_group.total_advance
                    exp.overdue_amount = (
                        used_total - group_total if used_total > group_total else Decimal("0.00")
                    )
                else:
                    exp.overdue_amount = Decimal("0.00")

        total_amount_to_be_paid = sum([exp.amount or Decimal("0.00") for exp in expenses])

        # ================== 🔥 LEADER ADVANCE BALANCE ==================
        if selected_user:
            leader_assignments = AdvanceAssignment.objects.filter(
                user=selected_user,
                is_leader=True,
                group__is_closed=False,   # ✅ only open groups
            )

            # 🔥 NEW DATE FILTER FOR GROUPS
            date_filter_groups = {}
            if from_date and to_date:
                try:
                    f = datetime.strptime(from_date, "%Y-%m-%d").date()
                    t = datetime.strptime(to_date, "%Y-%m-%d").date()
                    date_filter_groups = {
                        "created_at__date__gte": f,
                        "created_at__date__lte": t,
                    }
                except:
                    pass

            # 🔥 FILTER GROUPS BY CREATED_AT
            leader_assignments = leader_assignments.filter(group__in=AdvanceGroup.objects.filter(**date_filter_groups))

            for assignment in leader_assignments:
                group = assignment.group

                used_total = (
                    Expense.objects.filter(
                        advance_group=group,
                        is_allocation=False,
                        is_draft=False,
                        status__in=["pending", "approved", "paid"],
                        **date_filter_expenses
                    ).aggregate(total=Sum("amount"))["total"]
                    or Decimal("0.00")
                )

                raw_remaining = group.total_advance - used_total

                # ✅ Only POSITIVE remaining is treated as "advance balance"
                if raw_remaining > 0:
                    remaining = raw_remaining
                    overdue = Decimal("0.00")
                else:
                    remaining = Decimal("0.00")
                    overdue = -raw_remaining  # = used_total - total_advance

                leader_sub_balance += remaining
                leader_overdue_total += overdue

                leader_group_balances.append({
                    "group_name": group.name,
                    "created_at": group.created_at.date(),
                    "remaining": remaining,
                    "overdue": overdue
                })

            final_amount_to_pay = total_amount_to_be_paid - leader_sub_balance + leader_overdue_total

    # ---------------- CONTEXT -----------------
    context = {
        "users": users,
        "zones": zones,
        "expenses": expenses,
        "selected_user": selected_user,
        "selected_zone": selected_zone,
        "filtered_users": filtered_users,

        "total_amount_to_be_paid": total_amount_to_be_paid,
        "leader_sub_balance": leader_sub_balance,
        "leader_overdue_total": leader_overdue_total,
        "leader_group_balances": leader_group_balances,
        "final_amount_to_pay": final_amount_to_pay,

        "from_date": from_date,
        "to_date": to_date,
    }

    return render(request, "xp/payable.html", context)






# pay_now.py
from django.shortcuts import render
from .models import User, CashVoucher, Expense, AdvanceAssignment
from datetime import datetime
from decimal import Decimal

def pay_now(request, user_id):
    selected_user = User.objects.get(id=user_id)

    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    selected_filter = request.GET.get('selected_filter')


    # FIX → Always use ONLY the final_amount_to_pay coming from PARAM "amount"
    amount_param = request.GET.get('amount')
    try:
        total_amount_to_be_paid = Decimal(amount_param) if amount_param else Decimal("0.00")
    except:
        total_amount_to_be_paid = Decimal("0.00")

    # FIX → final_amount_to_pay MUST BE this same value, without recomputation here
    final_amount_to_pay = total_amount_to_be_paid

    # Leader advance balances (STILL NEEDED FOR DISPLAY PURPOSE ONLY)
    leader_assignments = AdvanceAssignment.objects.filter(user=selected_user, is_leader=True)
    total_leader_balance = sum([a.group.remaining_balance for a in leader_assignments]) or Decimal('0.00')

    # DO NOT CHANGE final_amount_to_pay HERE (just display)
    expenses = Expense.objects.filter(created_by=selected_user, status="pending")

    # Generate transaction ID
    current_year = str(datetime.now().year)[2:]
    current_month = str(datetime.now().month).zfill(2)
    counter = CashVoucher.objects.filter(created_by=selected_user).count() + 1
    transaction_id = f"XPTR{current_year}{current_month}{counter:04d}"
    queryset = Expense.objects.all()

    payment_details = None
    if request.method == 'POST':
        print("POST Data:", request.POST)
        transaction_id = request.POST.get('transaction_id')
        amount = request.POST.get('amount')
        from_date = request.POST.get('from_date')
        to_date = request.POST.get('to_date')
        screenshot = request.FILES.get('screenshot')
        remarks = request.POST.get('remarks', '').strip()

        payment_details = {
            'transaction_id': transaction_id,
            'amount': amount,
            'from_date': from_date,
            'to_date': to_date,
            'screenshot': screenshot,
            'remarks' : remarks,
        }

    return render(request, 'xp/pay_now.html', {
        'selected_user': selected_user,
        'expenses': expenses,
        'total_amount_to_be_paid': total_amount_to_be_paid,   # SAME as final_amount_to_pay
        'final_amount_to_pay': final_amount_to_pay,

        'transaction_id': transaction_id,
        'from_date': from_date,
        'to_date': to_date,
        'selected_filter': selected_filter,
        'payment_details': payment_details,
    })


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Sum
from decimal import Decimal
from datetime import datetime

from .models import Expense, Payment, User, Notification, AdvanceAssignment, AdvanceGroup

def process_payment(request):
    
    if request.method != "POST":
        messages.error(request, "Invalid request method.")
        return redirect('pay_now', user_id=request.POST.get('user_id'))

    try:
        user_id = request.POST.get('user_id')
        transaction_id = request.POST.get('transaction_id')
        amount_from_form = request.POST.get('amount') or "0"   # this is Final Amount to Pay shown in UI
        from_date = request.POST.get('from_date')
        to_date = request.POST.get('to_date')
        screenshot = request.FILES.get('screenshot')
        remarks = request.POST.get('remarks')

        selected_user = get_object_or_404(User, id=user_id)
        amount_from_form = Decimal(amount_from_form)

        from_date_obj = datetime.strptime(from_date, "%Y-%m-%d").date()
        to_date_obj = datetime.strptime(to_date, "%Y-%m-%d").date()

        print("=== PROCESS PAYMENT START ===")
        print(f"User: {selected_user.username}")
        print(f"Transaction ID: {transaction_id}")
        print(f"Final Amount from form (Pay Now): {amount_from_form}")
        print(f"From Date: {from_date_obj}  To Date: {to_date_obj}")

        # --- Get all pending expenses in date range ---
        expenses_query = Expense.objects.filter(
            created_by=selected_user,
            status="pending",
            transaction_date__range=[from_date_obj, to_date_obj]
        )
        total_expenses = expenses_query.aggregate(total=Sum('amount'))['total'] or Decimal(0)

        print(f"Pending Expenses Count: {expenses_query.count()}")
        print(f"Total Expenses (raw amount) in range: {total_expenses}")

        if total_expenses == 0:
            messages.warning(request, "No pending vouchers to pay.")
            return redirect('pay_now', user_id=user_id)

        # --- Leader groups & their balances ---
        leader_assignments = AdvanceAssignment.objects.filter(
            user=selected_user,
            is_leader=True
        ).select_related("group")

        total_leader_balance = sum(
            [max(a.group.remaining_balance, Decimal("0.00")) for a in leader_assignments]
        ) or Decimal("0.00")

        print(f"Total leader advance remaining (all groups): {total_leader_balance}")

        # --- How much advance can we apply in this payment? ---
        # We cannot use more advance than the voucher total
        advance_to_apply = min(max(total_leader_balance, 0), total_expenses)
        remaining_to_allocate = advance_to_apply

        print(f"Advance to apply in this payment: {advance_to_apply}")

        # ✅ Track which groups were actually used in THIS payment
        used_group_ids = set()

        # --- Deduct advance from each group until advance_to_apply is exhausted ---
        for assignment in leader_assignments:
            group = assignment.group

            if remaining_to_allocate <= 0:
                break  # we've used all advance we can use for this payment

            can_use = min(group.remaining_balance, remaining_to_allocate)

            if can_use <= 0:
                continue

            print(f"  Using ₹{can_use} from group '{group.name}' "
                  f"(before: remaining={group.remaining_balance}, used={group.used_advance})")

            group.used_advance += can_use
            group.remaining_balance -= can_use

            used_group_ids.add(group.id)  # ✅ mark this group as used

            if group.remaining_balance <= 0:
                group.remaining_balance = Decimal("0.00")
                group.is_closed = True
                print(f"  -> Group '{group.name}' fully used, closing group")

            group.save()

            remaining_to_allocate -= can_use

            print(f"  After: group '{group.name}' used_advance={group.used_advance}, "
                  f"remaining_balance={group.remaining_balance}")
            print(f"  Remaining advance to allocate: {remaining_to_allocate}")

        # --- Recompute final amount to pay based on what we actually deducted ---
        # Vouchers total - advance used in this payment
        # --------------------------
        # LEADER OVERDUE CALCULATION
        # --------------------------
        leader_overdue = Decimal("0.00")
        for assignment in leader_assignments:
            g = assignment.group
            # overdue = used_advance - total_advance
            diff = g.used_advance - g.total_advance
            if diff > 0:
                leader_overdue += diff

        print(f"Leader Overdue Total: {leader_overdue}")

        # --------------------------
        # CORRECT FINAL AMOUNT FORMULA (SAME AS UI)
        # --------------------------
        # total_leader_balance already calculated earlier
        computed_final_amount = total_expenses + leader_overdue - total_leader_balance

        if computed_final_amount < 0:
            computed_final_amount = Decimal("0.00")

        print(f"Final Amount to Pay (UI-MATCHED): {computed_final_amount}")

        print(f"Final amount to pay (computed): {computed_final_amount}")
        print(f"Final amount from form: {amount_from_form}")

        # Use the computed one (it should match the form; if not, logs will show)
        final_amount_to_pay = computed_final_amount

        # --- Mark all vouchers as PAID ---
        for expense in expenses_query:
            expense.status = "paid"
            expense.save()

            Notification.objects.create(
                user=expense.created_by,
                title="Voucher Paid",
                message=f"Voucher {expense.evoucher_number} of ₹{expense.amount} has been paid.",
            )

            print(f"Voucher {expense.id} marked as PAID")

        # -- CLEAR GROUP BALANCES AFTER PAYMENT --
        for assignment in leader_assignments:
            group = assignment.group

            # ✅ ONLY groups actually used in THIS payment
            if group.id not in used_group_ids:
                continue

            print(f"[CLEARING GROUP] {group.name}")

            if group.remaining_balance <= 0:
                group.remaining_balance = Decimal("0.00")

                # Clamp used_advance to total_advance → no more dues on this group
                if group.used_advance > group.total_advance:
                    group.used_advance = group.total_advance

                # ✅ Close group so it won't appear in future payable
                group.is_closed = True

                group.save(update_fields=["used_advance", "remaining_balance", "is_closed"])

                print(
                    f"  -> After clear: used={group.used_advance}, "
                    f"remaining={group.remaining_balance}, closed={group.is_closed}"
                )

        # --- Save Payment record ---
        Payment.objects.create(
            paid_to=selected_user,
            transaction_id=transaction_id,
            amount=final_amount_to_pay,
            from_date=from_date_obj,
            to_date=to_date_obj,
            screenshot=screenshot,
            remarks=remarks,
        )

        print(f"Payment Record Created: ₹{final_amount_to_pay}")
        print("=== PROCESS PAYMENT END ===")

        messages.success(
            request,
            f"Payment of ₹{final_amount_to_pay} processed successfully! "
            f"Vouchers marked paid and advances cleared."
        )
        return redirect('payable')

    except Exception as e:
        print(f"Error in process_payment: {str(e)}")
        messages.error(request, f"Error occurred: {str(e)}")
        return redirect('pay_now', user_id=request.POST.get('user_id'))
    


from django.shortcuts import render, redirect
from django.contrib import messages

def payment_success(request):
    # Add a success message
    messages.success(request, "Your payment has been successfully processed.")
    return redirect('payable')  # Redirect to the `payable` page or any relevant page






@login_required
def payments(request):
    print("entered Payments")
    query = request.GET.get('q', '')  # Search query
    user_id = request.GET.get('user_id')  # Get the user ID from the request
    
    user = request.user

    # Start with the base queryset depending on the user role
    if user.is_superuser:  # Admin - View all users' payments
        user_payments = Payment.objects.all().order_by('created_at')
    else:  # Regular user - View only own payments
        user_payments = Payment.objects.filter(paid_to=user).order_by('created_at')

    # If a user ID is selected, filter payments based on that user
    if user_id:
        user_payments = user_payments.filter(paid_to__id=user_id)

    # Apply search filters if a query is provided
    if query:
        user_payments = user_payments.filter(
            Q(transaction_id__icontains=query) |
            Q(paid_to__username__icontains=query) |
            Q(amount__icontains=query) |
            Q(from_date__icontains=query) |
            Q(to_date__icontains=query) |
            Q(remarks__icontains=query) 
        )

    # Prepare the list of payments with relevant details
    users = User.objects.all()  # Order users by username
    
    payments_with_details = []
    for payment in user_payments:
        payments_with_details.append({
            'id': payment.id,   # 🔥 Add this
            'transaction_id': payment.transaction_id,
            'paid_to': payment.paid_to,
            'amount': payment.amount,
            'from_date': payment.from_date,
            'to_date': payment.to_date,
            'remarks': payment.remarks, 
            'screenshot': payment.screenshot.url if payment.screenshot else None,
            'screenshot_name': payment.screenshot.name if payment.screenshot else "",  # 🔥 Add this
            'created_at': payment.created_at,
        })

    # Pass the filtered or unfiltered queryset with payment details and users to the template
    return render(request, 'xp/user_payment_history.html', {
        'payments': payments_with_details,
        'users': users,  # Add users here to be used in the template
        'selected_user_id': user_id,  # Pass the selected user ID back to the template
    })

@login_required
def update_payment(request, id):
    payment = get_object_or_404(Payment, id=id)

    if request.method == "POST":
        file = request.FILES.get("screenshot")
        remarks = request.POST.get("remarks", "").strip()

        if file:
            payment.screenshot = file

        # remarks should update EVEN if file not uploaded
        payment.remarks = remarks
        payment.save()

        messages.success(request, "Payment updated successfully!")

    return redirect("payments")

def get_filtered_payments(request):
    query = request.GET.get('q', '')  
    user_id = request.GET.get('user_id')
    from_date = request.GET.get('from')
    to_date = request.GET.get('to')

    payments = Payment.objects.all()

    if user_id:
        payments = payments.filter(paid_to_id=user_id)

    if from_date and to_date:
        try:
            from_dt = datetime.strptime(from_date, '%Y-%m-%d').date()
            to_dt = datetime.strptime(to_date, '%Y-%m-%d').date()
            payments = payments.filter(created_at__date__range=(from_dt, to_dt))
        except ValueError:
            pass

    if query:
        payments = payments.filter(
            Q(transaction_id__icontains=query) |
            Q(paid_to__username__icontains=query) |
            Q(amount__icontains=query)
        )

    return payments

import csv
from io import StringIO
from django.http import HttpResponse
from openpyxl import Workbook
from django.shortcuts import get_object_or_404
from django.utils.timezone import now
from reportlab.pdfgen import canvas

from .models import Payment  # Import your Payment model

def export_payments(request):
    file_format = request.GET.get('format', 'csv')
    payments = get_filtered_payments(request)

    if file_format == 'csv':
        return export_as_csv(payments, request)
    elif file_format == 'xlsx':
        return export_as_xlsx(payments, request)
    elif file_format == 'pdf':
        return export_as_pdf(request, payments)
    else:
        return HttpResponse("Invalid format requested", status=400)

# CSV Export
import csv
from django.http import HttpResponse
from django.utils.timezone import now
from django.conf import settings

def export_as_csv(payments, request):
    """
    Exports payment records as a CSV file with absolute file download links (if applicable).
    """
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="payments_{now().strftime("%Y%m%d")}.csv"'

    writer = csv.writer(response)
    writer.writerow(['Transaction ID', 'Paid To', 'Amount', 'From', 'To', 'Date', 'Uploaded Proof'])

    # Construct site URL for absolute file links
    site_url = request.build_absolute_uri('/')[:-1]  # Remove trailing slash

    for payment in payments:
        # Generate downloadable proof link
        if hasattr(payment, 'proof_photo') and payment.proof_photo:
            proof_file_url = f"{site_url}/download/{payment.proof_photo.name}"
        else:
            proof_file_url = 'No File'

        writer.writerow([
            payment.transaction_id or 'N/A',
            payment.paid_to or 'N/A',
            payment.amount or 0.00,
            payment.from_date.strftime('%Y-%m-%d') if payment.from_date else 'N/A',
            payment.to_date.strftime('%Y-%m-%d') if payment.to_date else 'N/A',
            payment.created_at.strftime('%d %b %Y, %H:%M') if payment.created_at else 'N/A',
            proof_file_url  # Downloadable link
        ])

    return response

# XLSX Export
from openpyxl import Workbook
from io import BytesIO
from django.http import HttpResponse
from django.utils.timezone import now

def export_as_xlsx(queryset, request):
    """
    Exports payment records as an Excel (.xlsx) file with absolute downloadable screenshot URLs.
    """
    # Create a new workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"

    # Define headers
    headers = ["Transaction ID", "Paid To", "Amount", "From Date", "To Date", "Screenshot URL", "Date"]
    ws.append(headers)

    # Construct site URL for absolute file links
    site_url = request.build_absolute_uri('/')[:-1]  # Remove trailing slash

    # Populate data rows
    for payment in queryset:
        # Generate downloadable screenshot link
        screenshot_url = f"{site_url}/download/{payment.screenshot.name}" if hasattr(payment, 'screenshot') and payment.screenshot else "No Screenshot"

        ws.append([
            payment.transaction_id or "N/A",
            str(payment.paid_to) if payment.paid_to else "N/A",
            payment.amount or 0.00,
            payment.from_date.strftime("%Y-%m-%d") if payment.from_date else "N/A",
            payment.to_date.strftime("%Y-%m-%d") if payment.to_date else "N/A",
            screenshot_url,
            payment.created_at.strftime("%Y-%m-%d %H:%M:%S") if payment.created_at else "N/A",
        ])

    # Save workbook to a BytesIO buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Return as an HTTP response
    response = HttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="payments_{now().strftime("%Y%m%d")}.xlsx"'

    return response

# PDF Export
from reportlab.pdfgen import canvas
from io import BytesIO
from django.http import HttpResponse

from io import BytesIO
from xhtml2pdf import pisa
from django.http import HttpResponse
from django.template.loader import render_to_string
from .models import Payment  # Assuming Payment is the model name, replace with actual model

def export_as_pdf(request, payments):
    """
    Export payment history to PDF using a styled HTML template.
    """
    # Fetch the payments you want to export; you can filter based on request parameters
    payments = Payment.objects.all()

    # Prepare context for the template
    context = {
        'payments': payments
    }

    # Render the HTML content using the template
    html_content = render_to_string('xp/export_payments_template.html', context)

    # Create an HttpResponse object with PDF content type
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="payments.pdf"'

    # Generate the PDF from HTML
    pdf_buffer = BytesIO()
    pisa_status = pisa.CreatePDF(html_content, dest=pdf_buffer)

    # Handle PDF generation errors
    if pisa_status.err:
        return HttpResponse('Error generating PDF', status=500)

    # Write the generated PDF to the response
    pdf_buffer.seek(0)
    response.write(pdf_buffer.read())
    return response
from django.http import JsonResponse

def get_monthwise_data(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    # Default to current month if no filters provided
    if not from_date or not to_date:
        today = date.today()
        first_day = today.replace(day=1)
        last_day = (today.replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        from_date = first_day
        to_date = last_day

    # Convert to date objects
    from_date = datetime.strptime(from_date, "%Y-%m-%d").date()
    to_date = datetime.strptime(to_date, "%Y-%m-%d").date()

    # Filter vouchers and expenses based on the date range
    vouchers = CashVoucher.objects.filter(date__range=(from_date, to_date))
    expenses = Expense.objects.filter(date__range=(from_date, to_date))

    # Get month-wise data
    vouchers_by_month = vouchers.annotate(month=TruncMonth('date')).values('month').annotate(
        total_claimed=Sum('amount'),
        total_paid=Sum('amount', filter=Q(status='paid'))
    ).order_by('month')

    expenses_by_month = expenses.annotate(month=TruncMonth('date')).values('month').annotate(
        total_claimed=Sum('amount'),
        total_paid=Sum('amount', filter=Q(status='paid'))
    ).order_by('month')

    months_list = []
    claimed_amounts = []
    paid_amounts = []

    for voucher in vouchers_by_month:
        months_list.append(voucher['month'].strftime('%b %Y'))
        claimed_amounts.append(voucher['total_claimed'])
        paid_amounts.append(voucher['total_paid'])

    for expense in expenses_by_month:
        if expense['month'].strftime('%b %Y') not in months_list:
            months_list.append(expense['month'].strftime('%b %Y'))
            claimed_amounts.append(expense['total_claimed'])
            paid_amounts.append(expense['total_paid'])

    return JsonResponse({
        'months_list': months_list,
        'claimed_amounts': claimed_amounts,
        'paid_amounts': paid_amounts,
    })


from django.shortcuts import render

def drafts_list(request):
    # Fetch draft items from the database (Example: filter by a 'draft' status)
    drafts = Expense.objects.filter(status='draft')
    return render(request, 'app/drafts_list.html', {'drafts': drafts})


# views.py
from django.http import JsonResponse
from .models import Conveyance

def get_price_per_km(request):
    vehicle_type = request.GET.get('vehicle_type')

    if not vehicle_type:
        return JsonResponse({'error': 'Vehicle type is required'}, status=400)

    try:
        # Fetch the conveyance object based on the vehicle type
        conveyance = Conveyance.objects.get(vehicle_type=vehicle_type)
        return JsonResponse({'price_per_km': float(conveyance.price_per_km)})
    except Conveyance.DoesNotExist:
        return JsonResponse({'error': 'Price per kilometer not found for this vehicle type'}, status=404)
    

from django.views.decorators.http import require_POST

@require_POST
def reject_cash_voucher_ajax(request):
    print("GOT IN")
    voucher_id = request.POST.get('voucher_id')
    reason = request.POST.get('reason', '')
    
    if not voucher_id or not reason:
        return JsonResponse({'error': 'Missing parameters'}, status=400)

    try:
        voucher = CashVoucher.objects.get(id=voucher_id)
        voucher.status = 'rejected'
        voucher.rejection_reason = reason
        voucher.save()
        return JsonResponse({'message': 'Voucher rejected successfully!'})
    except CashVoucher.DoesNotExist:
        return JsonResponse({'error': 'Voucher not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    


from django.shortcuts import render

def service_feedback_form(request):
    return render(request, 'xp/feedbackform.html') 



from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from django.utils import timezone

from django.http import JsonResponse

@require_POST
@login_required
def reject_expense(request):
    expense_id = request.POST.get('expense_id')
    reason = request.POST.get('reason')

    if not expense_id or not reason:
        return JsonResponse({'success': False, 'error': 'Missing data'}, status=400)

    try:
        expense = Expense.objects.get(id=expense_id)
        expense.is_rejected = True
        expense.rejection_reason = reason
        expense.rejected_at = timezone.now()
        expense.is_approved = False  # optional: reset approval
        # DO NOT TOUCH status here
        expense.save()

        return JsonResponse({'success': True})
    except Expense.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Expense not found'}, status=404)

from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from website.models import Expense

@require_POST
@login_required
def approve_expense(request, expense_id):
    try:
        expense = Expense.objects.get(id=expense_id)
        if expense.cash_vouchers.filter(voucher_number__startswith="XPCV", status="pending").exists():
            return JsonResponse({'success': False, 'error': 'Please approve/reject the tip first'}, status=400)

        expense.is_approved = True
        expense.is_rejected = False
        expense.rejection_reason = None
        expense.rejected_at = None

        # ✅ Mark as PAID if it's an advance
        if expense.item_type == 'advances':
            expense.status = 'paid'

        expense.save()

        return JsonResponse({'success': True})
    
    except Expense.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Expense not found'}, status=404)
    

# xp/views.py

from django.shortcuts import render, redirect
from .models import PaymentRequest
from django.contrib.auth.decorators import login_required
from django.contrib import messages

@login_required

def submit_request(request):
    if request.method == 'POST':
        message = request.POST.get('message')
        if not message:
            messages.error(request, "Message cannot be empty.")
            return redirect('submit_request')

        req = PaymentRequest.objects.create(user=request.user)

        PaymentRequestMessage.objects.create(
            request=req,
            sender=request.user,
            message=message
        )

        messages.success(request, "Request submitted successfully.")
        return redirect('view_requests')

    return render(request, 'xp/submit_request.html')

# xp/views.py

from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from .models import PaymentRequest

@login_required
def view_requests(request):
    if request.user.is_superuser or request.user.is_staff:
        requests = PaymentRequest.objects.all().order_by('-created_at')  # Admin sees all
    else:
        requests = PaymentRequest.objects.filter(user=request.user).order_by('-created_at')  # User sees only their own
    return render(request, 'xp/view_requests.html', {'requests': requests})


from django.shortcuts import get_object_or_404, redirect, render
from django.http import JsonResponse
from .models import PaymentRequest, PaymentRequestMessage
from django.utils.timezone import localtime

def chat_request(request, req_id):
    req = get_object_or_404(PaymentRequest, id=req_id)

    if request.method == "POST":
        msg = request.POST.get("message", "").strip()
        mark_resolved = request.POST.get("resolve")
        is_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"

        response = {"success": True}

        if msg:
            message_obj = PaymentRequestMessage.objects.create(
                request=req,
                sender=request.user,
                message=msg
            )
            response.update({
                "sender": message_obj.sender.username,
                "message": message_obj.message,
                "sent_at": localtime(message_obj.sent_at).strftime("%H:%M"),
            })

        # If resolved checkbox is checked
        if mark_resolved:
            req.is_resolved = True
            req.save()
            response["resolved"] = True

            # Include sender even if no message was sent (to avoid "undefined")
            if "sender" not in response:
                response["sender"] = request.user.username

        if is_ajax:
            return JsonResponse(response)
        else:
            return redirect('chat_request', req_id=req.id)

    return render(request, 'xp/chat_request.html', {'req': req})

# views.py

from django.http import JsonResponse
from .models import PaymentRequest, PaymentRequestMessage
from django.shortcuts import get_object_or_404
from django.utils.timezone import now

def ajax_reply_request(request):
    if request.method == 'POST':
        req_id = request.POST.get('request_id')
        msg = request.POST.get('message', '').strip()
        mark_resolved = request.POST.get('resolve')
        file = request.FILES.get('attachment')
        req = get_object_or_404(PaymentRequest, id=req_id)

        response = {"success": True}

        # Create message if message or file is present
        if msg or file:
            message_obj = PaymentRequestMessage.objects.create(
                request=req,
                sender=request.user,
                message=msg,
                attachment=file if file else None
            )
            response.update({
                "sender": message_obj.sender.username,
                "message": message_obj.message,
                "sent_at": message_obj.sent_at.strftime("%H:%M"),
                "attachment_url": message_obj.attachment.url if message_obj.attachment else None,
                "attachment_name": message_obj.attachment.name if message_obj.attachment else None
            })

        # Mark as resolved
        if mark_resolved:
            req.is_resolved = True
            req.save()
            response["resolved"] = True

        return JsonResponse(response)

from django.http import JsonResponse
from .models import PaymentRequest, PaymentRequestMessage

from django.contrib.auth.decorators import login_required
from django.utils.timesince import timesince

from django.http import JsonResponse
from .models import PaymentRequest, PaymentRequestMessage
from django.utils.timesince import timesince
from django.contrib.auth.decorators import login_required

@login_required
def fetch_messages(request, req_id):
    try:
        req = PaymentRequest.objects.get(id=req_id)
    except PaymentRequest.DoesNotExist:
        return JsonResponse({"success": False, "error": "Request not found"}, status=404)

    messages = req.messages.select_related("sender").order_by("sent_at")
    result = []
    for msg in messages:
        result.append({
            "sender": msg.sender.username,
            "message": msg.message,
            "timestamp": timesince(msg.sent_at) + " ago",
            "attachment": msg.attachment.url if msg.attachment else None,
            "attachment_name": msg.attachment.name.split("/")[-1] if msg.attachment else "",
            "is_me": request.user == msg.sender
        })

    return JsonResponse({"success": True, "messages": result})


from django.views.decorators.http import require_POST
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import Expense

@require_POST
@login_required
def request_expense_deletion(request, expense_id):
    expense = get_object_or_404(Expense, id=expense_id)
    if not expense.delete_requested:
        expense.delete_requested = True
        expense.delete_requested_by = request.user
        expense.save()
        if request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"success": True, "requested": True})
        messages.success(request, "📝 Deletion request submitted.")
    else:
        if request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"success": False, "error": "Already requested"})
        messages.info(request, "Already requested.")
    return redirect("item_form")


@require_POST
@login_required
def delete_expense(request, expense_id):
    if not request.user.is_staff and not request.user.is_superuser:
        if request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"success": False, "error": "Not authorized"})
        messages.error(request, "You are not authorized.")
        return redirect("item_form")

    expense = get_object_or_404(Expense, id=expense_id, delete_requested=True)
    expense.delete()
    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return JsonResponse({"success": True, "deleted": True})
    messages.success(request, "✅ Expense deleted.")
    return redirect("item_form")


from django.contrib.auth import authenticate
from django.contrib import messages
from django.shortcuts import redirect, get_object_or_404
from .models import PaymentRequest

from django.views.decorators.csrf import csrf_exempt  # Optional

@csrf_exempt  # Optional: only if your CSRF isn't working
def delete_request(request):
    if request.method == 'POST':
        request_id = request.POST.get('request_id')
        admin_password = request.POST.get('admin_password')

        if not request.user.is_superuser:
            messages.error(request, "Only admins can delete requests.")
            return redirect('view_requests')

        # Authenticate admin again
        user = authenticate(username=request.user.username, password=admin_password)
        if user is None:
            messages.error(request, "Invalid admin password.")
            return redirect('view_requests')

        # Delete the request
        payment_request = get_object_or_404(PaymentRequest, id=request_id)
        payment_request.delete()
        messages.success(request, "Request deleted successfully.")
        return redirect('view_requests')

    return redirect('view_requests')

from django.db.models import Sum
from decimal import Decimal
from django.shortcuts import get_object_or_404
from django.http import JsonResponse

def get_group_balance(request, group_id):
    group = get_object_or_404(AdvanceGroup, id=group_id)
    
    # 🚨 Only count real expenses, not allocations
    used = (
        Expense.objects.filter(
            item_type="advances",
            advance_group=group,
            is_allocation=False   # <-- critical fix
        )
        .aggregate(total=Sum("amount"))["total"]
        or Decimal("0.00")
    )

    remaining = group.total_advance - used
    return JsonResponse({"remaining_balance": float(remaining)})


from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from .models import AdvanceGroup
from django.contrib.auth.decorators import login_required, user_passes_test

# Only staff or superuser can close
def staff_required(user):
    return user.is_staff or user.is_superuser

@login_required
@user_passes_test(staff_required)
def close_group(request, group_id):
    group = get_object_or_404(AdvanceGroup, id=group_id)
    if group.is_closed:
        messages.warning(request, f"Group '{group.name}' is already closed.")
    else:
        group.is_closed = True
        group.save()
        messages.success(request, f"Group '{group.name}' has been closed successfully.")
    return redirect('view_assigned_groups')  # replace with your template/view



from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils.crypto import get_random_string
from datetime import date
from .models import DirectPay, Expense



# Form to create a DirectPay record
def directpay_form(request):
    if request.method == "POST":
        dp = DirectPay.objects.create(
            created_by=request.user,
            item_type=request.POST.get("item_type"),
            item_name=request.POST.get("item_name"),
            payment_category=request.POST.get("transaction_option"),
            transaction_category=request.POST.get("transaction_category"),
            external_name=request.POST.get("external_name"),
            total_amount=request.POST.get("total_amount"),
            payment_mode=request.POST.get("payment_mode"),
            proof_photo=request.FILES.get("proof_photo"),
            bill_file=request.FILES.get("bill_photo"),  # matches model field 'bill_file'
            gst_file=request.FILES.get("gst_photo"),    # matches model field 'gst_file'
            transaction_date=request.POST.get("transaction_date"),
        )
        print("bill file:", request.FILES.get("bill_photo"))
        print("gst file:", request.FILES.get("gst_photo"))
        return redirect("directpay_preview")

    context = {

        "current_date": date.today(),
    }
    return render(request, "xp/directpay_form.html", context)


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.models import User
from django.db.models import Q
from datetime import datetime
from django.http import JsonResponse
from django.template.loader import render_to_string

from .models import DirectPay, Expense


def directpay_preview(request):
    # Base query
    if request.user.is_staff:
        directpays = DirectPay.objects.all()
    else:
        directpays = DirectPay.objects.filter(created_by=request.user)

    # Filters
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    sort_order = request.GET.get("sort_order", "desc")
    search_query = request.GET.get("q")
    user_id = request.GET.get("user_id")

    if user_id:
        directpays = directpays.filter(created_by_id=user_id)

    if search_query:
        directpays = directpays.filter(
            Q(item_name__icontains=search_query) |
            Q(external_name__icontains=search_query)
        )

    if start_date and end_date:
        directpays = directpays.filter(transaction_date__range=[start_date, end_date])

    directpays = directpays.order_by("transaction_date" if sort_order == "asc" else "-transaction_date")

    # ---------------------------
    # 🔥 AJAX Voucher Creation
    # ---------------------------
    if request.method == "POST" and request.headers.get("x-requested-with") == "XMLHttpRequest":
        selected_ids = request.POST.getlist("selected_ids[]", [])
        single_dp_id = request.POST.get("single_dp_id")

        if single_dp_id:
            selected_ids = [single_dp_id]

        if not selected_ids:
            return JsonResponse({"success": False, "message": "Please select at least one record."})

        target_user_id = request.POST.get("create_for_user")
        if target_user_id:
            system_user = get_object_or_404(User, id=target_user_id)
        else:
            system_user = User.objects.get(id=16)  # fallback default
        created_count = 0

        for dp_id in selected_ids:
            dp = get_object_or_404(DirectPay, id=dp_id)
            new_evoucher = generate_evoucher_number()

            Expense.objects.create(
                created_by=system_user,
                item_type=dp.item_type,
                item_name=dp.item_name,
                transaction_option=dp.payment_category,
                transaction_category=dp.transaction_category,
                external_type=dp.external_name,
                amount=dp.total_amount,
                payment_mode=dp.payment_mode,
                evoucher_number=new_evoucher,
                proof_photo=dp.proof_photo,
                bill_photo=getattr(dp, "bill_file", None),
                gst_photo=getattr(dp, "gst_file", None),
                transaction_date=dp.transaction_date,
                status="pending",
                is_approved=True,
                is_rejected=False,
            )

            dp.delete()
            created_count += 1

        # Re-render table with filters intact
        html = render_to_string(
            "xp/directpay_preview.html",
            {
                "directpays": directpays,
                "start_date": start_date,
                "end_date": end_date,
                "sort_order": sort_order,
                "all_users": User.objects.all().order_by("first_name"),
                "selected_user_id": int(user_id) if user_id else None,
                "search_query": search_query or "",
            },
            request=request,
        )

        return JsonResponse({
            "success": True,
            "created": created_count,
            "html": html
        })

    # ---------------------------
    # Normal POST (non-AJAX)
    # ---------------------------
    if request.method == "POST":
        selected_ids = request.POST.getlist("selected_ids")
        single_dp_id = request.POST.get("single_dp_id")

        if single_dp_id:
            selected_ids = [single_dp_id]

        if not selected_ids:
            messages.warning(request, "⚠️ Please select at least one record.")
            return redirect("directpay_preview")

        target_user_id = request.POST.get("create_for_user")
        if target_user_id:
            system_user = get_object_or_404(User, id=target_user_id)
        else:
            system_user = User.objects.get(id=16)  # fallback default
        created_count = 0

        for dp_id in selected_ids:
            dp = get_object_or_404(DirectPay, id=dp_id)
            new_evoucher = generate_evoucher_number()

            Expense.objects.create(
                created_by=system_user,
                item_type=dp.item_type,
                item_name=dp.item_name,
                transaction_option=dp.payment_category,
                transaction_category=dp.transaction_category,
                external_type=dp.external_name,
                amount=dp.total_amount,
                payment_mode=dp.payment_mode,
                evoucher_number=new_evoucher,
                proof_photo=dp.proof_photo,
                bill_photo=getattr(dp, "bill_file", None),
                gst_photo=getattr(dp, "gst_file", None),
                transaction_date=dp.transaction_date,
                status="pending",
                is_approved=True,
                is_rejected=False,
            )

            dp.delete()
            created_count += 1

        messages.success(request, f"✅ {created_count} voucher(s) created successfully.")
        return redirect("directpay_preview")

    # ---------------------------
    # Render Page
    # ---------------------------
    all_users = User.objects.all().order_by("first_name")
    context = {
        "directpays": directpays,
        "start_date": start_date,
        "end_date": end_date,
        "sort_order": sort_order,
        "all_users": all_users,
        "selected_user_id": int(user_id) if user_id else None,
        "search_query": search_query or "",
    }
    return render(request, "xp/directpay_preview.html", context)  

def edit_directpay_form(request, dp_id):
    dp = get_object_or_404(DirectPay, id=dp_id)

    if request.method == "POST":
        # Update fields
        dp.item_type = request.POST.get("item_type")
        dp.item_name = request.POST.get("item_name")
        dp.payment_category = request.POST.get("transaction_option")
        dp.transaction_category = request.POST.get("transaction_category")
        dp.external_name = request.POST.get("external_name")
        dp.total_amount = request.POST.get("total_amount")
        dp.payment_mode = request.POST.get("payment_mode")
        dp.transaction_date = request.POST.get("transaction_date")

        # Only update files if new files are uploaded
        if request.FILES.get("proof_photo"):
            dp.proof_photo = request.FILES.get("proof_photo")
        if request.FILES.get("bill_photo"):
            dp.bill_file = request.FILES.get("bill_photo")
        if request.FILES.get("gst_photo"):
            dp.gst_file = request.FILES.get("gst_photo")

        dp.save()
        messages.success(request, "DirectPay updated successfully.")
        return redirect("directpay_preview")

    # Prefill form with existing DirectPay instance
    context = {
        "dp": dp,
        "current_date": dp.transaction_date or date.today(),
    }
    return render(request, "xp/edit_directpay_form.html", context)

from django.urls import reverse

def delete_advance_log(request, log_id):
    log = get_object_or_404(AdvanceGroupUpdateLog, id=log_id)

    # Optional: delete stored file
    if log.proof_file and hasattr(log.proof_file, 'path'):
        try:
            import os

            if os.path.exists(log.proof_file.path):
                os.remove(log.proof_file.path)
        except Exception:
            logger.exception("Failed to delete proof file from disk")

    group_name = log.group.name
    log.delete()

    messages.success(request, "History record deleted successfully.")
    return redirect(reverse("assign_advances") + f"?group_name={group_name}&action=history")



from django.shortcuts import render, redirect
from django.db.models import Sum
from decimal import Decimal
from .models import AdvanceGroup, Expense
from django.contrib import messages

def fix_advance_balances(request):
    """
    STEP 1: PREVIEW PAGE
    Shows OLD and NEW calculated values but does NOT update DB
    """

    preview_data = []

    groups = AdvanceGroup.objects.all()

    for group in groups:
        old_used = group.used_advance or Decimal("0")
        old_remaining = group.remaining_balance or Decimal("0")

        # 🔥 TRUE used amount from Expenses (correct logic)
        new_used = (
            Expense.objects.filter(
                advance_group=group,
                is_allocation=False,
                is_draft=False,
                status__in=["pending", "approved", "paid"]
            ).aggregate(total=Sum("amount"))["total"]
            or Decimal("0")
        )

        new_remaining = group.total_advance - new_used


        preview_data.append({
            "group": group,
            "old_used": old_used,
            "old_remaining": old_remaining,
            "new_used": new_used,
            "new_remaining": new_remaining,
        })

    return render(request, "xp/fix_advance_balances.html", {
        "preview_data": preview_data
    })


def confirm_fix_advance_balances(request):
    """
    STEP 2: ACTUALLY WRITE FIXED VALUES INTO DATABASE
    """

    groups = AdvanceGroup.objects.all()

    for group in groups:
        new_used = (
            Expense.objects.filter(
                advance_group=group,
                is_allocation=False,
                is_draft=False,
                status__in=["pending", "approved", "paid"]
            ).aggregate(total=Sum("amount"))["total"]
            or Decimal("0")
        )

        new_remaining = group.total_advance - new_used
  

        group.used_advance = new_used
        group.remaining_balance = new_remaining
        group.save()

    messages.success(request, "✅ Successfully updated all AdvanceGroup balances!")
    return redirect("fix_advance_balances")



# website/views.py
import secrets
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import redirect
from django.contrib.auth import login

from .models import MobileSession, FCMToken


@login_required
def mobile_session(request):
    session, _ = MobileSession.objects.get_or_create(
        user=request.user,
        defaults={"token": secrets.token_hex(32)}
    )
    return JsonResponse({"token": session.token})


@csrf_exempt
def mobile_token_login(request):
    token = request.headers.get("X-MOBILE-TOKEN")

    if not token:
        return redirect("/riva/login/")

    try:
        session = MobileSession.objects.get(token=token)
        user = session.user
        login(request, user)

        MobileSession.objects.get_or_create(
            user=user,
            defaults={"token": token}
        )

        return redirect(f"/website/home?mobile_token={session.token}")

    except MobileSession.DoesNotExist:
        return redirect("/riva/login/")


@csrf_exempt
def save_fcm_token(request):
    print("🔥 SAVE FCM TOKEN HIT")

    mobile_token = request.headers.get("Authorization")
    if not mobile_token:
        return JsonResponse({"error": "no auth"}, status=401)

    mobile_token = mobile_token.replace("Bearer ", "")

    try:
        session = MobileSession.objects.get(token=mobile_token)
        user = session.user
    except MobileSession.DoesNotExist:
        return JsonResponse({"error": "invalid token"}, status=401)

    fcm_token = request.POST.get("fcm_token")
    print("🔥 USER:", user)
    print("🔥 FCM TOKEN:", fcm_token)

    if fcm_token:
        obj, _ = FCMToken.objects.get_or_create(
            user=user,
            token=fcm_token
        )
        print("🔥 SAVED:", obj.token)

    return JsonResponse({"status": "ok"})

